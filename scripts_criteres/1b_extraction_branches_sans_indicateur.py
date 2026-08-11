# -*- coding: utf-8 -*-
"""
ÉTAPE 1b — Extraction des 4 branches ajoutées après coup (cible seule).

Rôle : Administration publique, Éducation-santé, Services aux entreprises
et Autres services ont été ajoutées au classeur après les 12 branches
principales, avec une mise en page plus simple (uniquement des variantes
de cible, aucun indicateur). Ce script les extrait séparément avec une
règle plus simple que le script 1a : la colonne de dates associée à une
cible est toujours celle qui la précède immédiatement (pas de recherche
en arrière sur plusieurs colonnes).

Entrée  : Etude_sectorielle_Maroc_2_complete.xlsx (4 feuilles concernées)
Sortie  : extra_branches.pkl ({branche: liste de tuples (date, valeur)})

Dépendances : openpyxl
    pip install openpyxl
"""
import re
from datetime import date
from openpyxl import load_workbook
import pickle

SRC = "Etude_sectorielle_Maroc_2_complete.xlsx"
OUT = "extra_branches.pkl"

EXTRA_BRANCHES = ["Administration publique", "Éducation-santé",
                   "Services aux entreprises", "Autres services"]


def parse_date_label(lbl):
    """Reconnaît 'T1-2014' et '2014T1' (les deux ordres rencontrés selon
    les feuilles) — voir script 1a pour la liste complète des formats."""
    if lbl is None:
        return None
    s = str(lbl).strip()
    m = re.match(r"^T([1-4])-(\d{4})$", s)
    if m:
        t, y = int(m.group(1)), int(m.group(2))
        return date(y, (t - 1) * 3 + 1, 1)
    m = re.match(r"^(\d{4})T([1-4])$", s)
    if m:
        y, t = int(m.group(1)), int(m.group(2))
        return date(y, (t - 1) * 3 + 1, 1)
    return None


if __name__ == "__main__":
    wb = load_workbook(SRC, data_only=True)
    resultats = {}

    for b in EXTRA_BRANCHES:
        ws = wb[b]
        max_col, max_row = ws.max_column, ws.max_row
        best = None
        for c in range(1, max_col + 1):
            nom = ws.cell(row=3, column=c).value
            if not nom or not isinstance(nom, str):
                continue
            # on ne garde que les colonnes qui sont des variantes de la
            # cible (VA de branche), reperees par leur nom
            if not re.search(r"\bVA\b|rétropolée", nom):
                continue
            # la colonne de dates est la derniere colonne avant c dont la
            # ligne 5 est un libelle de date reconnu
            date_col = None
            for cc in range(c - 1, 0, -1):
                v = ws.cell(row=5, column=cc).value
                if isinstance(v, str) and parse_date_label(v) is not None:
                    date_col = cc
                    break
            if date_col is None:
                continue
            serie = []
            for r in range(5, max_row + 1):
                dlabel = ws.cell(row=r, column=date_col).value
                dt = parse_date_label(dlabel)
                v = ws.cell(row=r, column=c).value
                if dt is not None and isinstance(v, (int, float)):
                    serie.append((dt, float(v)))
            # on garde la plus longue serie candidate (typiquement la
            # version "retropolee sur 28 branches", la plus longue)
            if len(serie) >= 100 and (best is None or len(serie) > len(best)):
                best = serie

        resultats[b] = best
        if best:
            s = sorted(best)
            print(f"{b:28s} -> {len(s)} obs, {s[0][0]} à {s[-1][0]}")
        else:
            print(f"{b:28s} -> aucune série exploitable trouvée")

    with open(OUT, "wb") as f:
        pickle.dump(resultats, f)
    print(f"\nDonnées sauvegardées dans {OUT}")
