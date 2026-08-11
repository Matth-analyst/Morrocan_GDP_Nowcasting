# -*- coding: utf-8 -*-
"""
ÉTAPE 1a — Extraction des données brutes depuis le classeur consolidé.

Rôle : lit chaque feuille de branche du classeur Excel et en extrait,
sans aucun filtre, la variable cible et tous les indicateurs candidats,
avec leurs dates réelles associées.

Entrée  : Etude_sectorielle_Maroc_2_complete.xlsx (feuilles par branche)
Sortie  : parsed_branches_v2.pkl (dictionnaire {branche: {cible, indicateurs, cibles_add}})

Dépendances : openpyxl, numpy, pandas
    pip install openpyxl numpy pandas

MÉTHODE DE PARSING (importante pour comprendre le résultat) :
Chaque feuille de branche mélange plusieurs blocs (variable cible,
indicateurs mensuels, indicateurs trimestriels, parfois plusieurs blocs
successifs ajoutés au fil de la collecte). Une première version de ce
script (v1) déterminait le type d'une colonne (date ou indicateur) en
regardant uniquement sa 5e ligne — ce qui ratait des centaines de
colonnes dont la première valeur était vide. Cette version (v2) corrige
le problème : on regarde le TYPE DOMINANT (date vs numérique) de TOUTES
les valeurs non vides de la colonne avant de la classer, puis chaque
indicateur est rattaché à la DERNIÈRE colonne de dates rencontrée avant
lui dans l'ordre des colonnes.

LIMITE CONNUE : le champ "freq" renvoyé pour chaque indicateur n'est pas
renseigné à cette étape (toujours None) — la fréquence réelle est
recalculée à partir des dates elles-mêmes dans le script 2 (méthode plus
fiable qu'une étiquette de colonne, qui peut être absente ou ambiguë).
"""
import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from datetime import date
import pickle

SRC = "Etude_sectorielle_Maroc_2_complete.xlsx"
OUT = "parsed_branches_v2.pkl"

# Les 12 branches traitées par cette version du script. Les 4 branches
# ajoutées ensuite (Administration publique, Éducation-santé, Services
# aux entreprises, Autres services) n'ont qu'une cible et pas
# d'indicateur : elles sont traitées séparément par le script 1b.
BRANCH_SHEETS = ["Agriculture", "Pêche", "Industrie d'extraction",
    "Industrie de transformation", "Électricité, gaz, eau", "Construction",
    "Commerce", "Transports", "Hébergement-restauration",
    "Information-communication", "Finances et assurances", "Immobilier"]

MOIS_FR = {"janv": 1, "févr": 2, "fev": 2, "mars": 3, "avr": 4, "mai": 5,
           "juin": 6, "juil": 7, "août": 8, "aout": 8, "sept": 9, "oct": 10,
           "nov": 11, "déc": 12, "dec": 12}


def parse_date_label(lbl):
    """Reconnaît les formats de date rencontrés dans le classeur et
    renvoie (date_python, frequence_deduite_du_libelle) ou (None, None)
    si le libellé n'est pas une date reconnue.
    Formats gérés : 'T1-2014', '2014T1', 'janv-14', '2014M01', '2014'."""
    if lbl is None:
        return None, None
    s = str(lbl).strip()

    m = re.match(r"^T([1-4])-(\d{4})$", s)
    if m:
        t, y = int(m.group(1)), int(m.group(2))
        return date(y, (t - 1) * 3 + 1, 1), "trimestriel"

    m = re.match(r"^(\d{4})T([1-4])$", s)
    if m:
        y, t = int(m.group(1)), int(m.group(2))
        return date(y, (t - 1) * 3 + 1, 1), "trimestriel"

    m = re.match(r"^([a-zéû]+)-(\d{2})$", s.lower())
    if m and m.group(1) in MOIS_FR:
        mois = MOIS_FR[m.group(1)]
        yy = int(m.group(2))
        year = 2000 + yy if yy < 70 else 1900 + yy
        return date(year, mois, 1), "mensuel"

    m = re.match(r"^(\d{4})M(\d{1,2})$", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), 1), "mensuel"

    m = re.match(r"^(\d{4})$", s)
    if m:
        return date(int(m.group(1)), 1, 1), "annuel"

    return None, None


def column_values(ws, c, max_row):
    """Renvoie {numéro_de_ligne: valeur} pour toutes les cellules non
    vides d'une colonne, à partir de la ligne 5 (les 4 premières lignes
    sont réservées aux titres/en-têtes de section)."""
    vals = {}
    for r in range(5, max_row + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            vals[r] = v
    return vals


def classify_column(vals):
    """Classe une colonne comme 'date' (majorité de libellés de date),
    'indicateur' (majorité de valeurs numériques) ou 'vide'."""
    if not vals:
        return "vide"
    n_date, n_num = 0, 0
    for v in vals.values():
        if isinstance(v, str):
            dt, _ = parse_date_label(v)
            if dt is not None:
                n_date += 1
        elif isinstance(v, (int, float)):
            n_num += 1
    if n_date >= n_num and n_date > 0:
        return "date"
    if n_num > 0:
        return "indicateur"
    return "vide"


def load_branch_sheet(ws):
    """Extrait de la feuille : la cible standard (colonnes A/B), et pour
    chaque colonne à partir de D, si elle contient un indicateur, sa
    série (date, valeur) correctement associée à sa colonne de dates."""
    max_col, max_row = ws.max_column, ws.max_row

    # --- Cible standard (colonnes A/B) ---
    cible_serie = []
    for r in range(5, max_row + 1):
        d = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        dt, _ = parse_date_label(d)
        if dt is not None and isinstance(v, (int, float)):
            cible_serie.append((dt, float(v)))

    # --- Classification de toutes les colonnes à partir de D ---
    col_info = {}
    for c in range(4, max_col + 1):
        vals = column_values(ws, c, max_row)
        col_info[c] = dict(type=classify_column(vals), vals=vals,
                            nom=ws.cell(row=3, column=c).value,
                            source=ws.cell(row=4, column=c).value)

    # --- Rattachement de chaque indicateur à sa colonne de dates ---
    indicateurs, cibles_additionnelles = [], []
    current_date_col = None
    for c in range(4, max_col + 1):
        info = col_info[c]
        if info["type"] == "date":
            current_date_col = c
            continue
        if info["type"] != "indicateur" or current_date_col is None:
            continue
        nom = info["nom"]
        if not nom:
            continue
        date_vals = col_info[current_date_col]["vals"]
        serie = []
        for r, v in info["vals"].items():
            dlabel = date_vals.get(r)
            dt, freq = parse_date_label(dlabel) if dlabel is not None else (None, None)
            if dt is not None and isinstance(v, (int, float)):
                serie.append((dt, float(v)))
        if not serie:
            continue
        entry = dict(nom=str(nom), source=info["source"], serie=serie, freq=None)
        # Heuristique : un nom contenant "VA", "PIB trimestriel" ou une
        # mention de base comptable est une variante de cible potentielle,
        # pas un indicateur d'activité.
        if re.search(r"\bVA\b|PIB trimestriel|base 2014|base 2007", str(nom)):
            cibles_additionnelles.append(entry)
        else:
            indicateurs.append(entry)
    return cible_serie, indicateurs, cibles_additionnelles


if __name__ == "__main__":
    wb = load_workbook(SRC, data_only=True)
    all_data = {}
    for b in BRANCH_SHEETS:
        ws = wb[b]
        cible, indicateurs, cibles_add = load_branch_sheet(ws)
        all_data[b] = dict(cible=cible, indicateurs=indicateurs, cibles_add=cibles_add)
        print(f"{b:32s} cible_base={len(cible):3d} obs | "
              f"indicateurs={len(indicateurs):3d} | cibles_alt_trouvees={len(cibles_add):3d}")

    with open(OUT, "wb") as f:
        pickle.dump(all_data, f)
    print(f"\nDonnées sauvegardées dans {OUT}")
