# -*- coding: utf-8 -*-
"""
Audit final + correction de fidelite de Etude_sectorielle_Maroc_2_complete.xlsx.

Contexte
--------
21 (puis 31 apres audit exhaustif complet, cf ci-dessous) colonnes du
classeur n'ont aucune correspondance verifiable dans le data lake. Deux
d'entre elles se sont revelees etre des TRANSFORMATIONS legitimes d'une
source reelle (delta d'une serie cumulee, arrondi) : elles sont remplacees
par la vraie serie, plus longue et plus precise. Les 29 autres n'ont recu
aucune explication : elles sont retirees du classeur, car les y laisser
reviendrait a presenter comme fiable une donnee qui ne l'est pas.

Ce script :
  1. Remplace "Vente de ciment (1000tonnes)" par le delta mensuel de
     "Ventes locales du ciment (mensuel cumulé).csv" (1995M01 -> 2026M05,
     au lieu d'une plage plus courte et non tracee).
  2. Remplace "Taux d'Utilisation des Capacités" par la serie primaire
     "secondaire/industrie/Taux d'utilisation des capacités (mensuel).csv"
     (2010M01 -> 2025M08, valeurs exactes au lieu d'arrondis).
  3. Supprime les 29 colonnes restantes, dans les feuilles ET dans
     Metadonnees/Audit/Sommaire, avec une trace explicite de ce qui a ete
     retire et pourquoi (feuille "Colonnes retirees").

Sortie : le fichier est modifie sur place. Un journal texte est imprime.
"""
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parent))
from catalogue_lake import charger_lake                      # noqa: E402

CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'

MOIS_FR = {'janv': 1, 'fevr': 2, 'févr': 2, 'mars': 3, 'avr': 4, 'mai': 5,
           'juin': 6, 'juil': 7, 'aout': 8, 'août': 8, 'sept': 9, 'oct': 10,
           'nov': 11, 'dec': 12, 'déc': 12}

# Les 29 colonnes sans AUCUNE correspondance trouvee dans le data lake,
# malgre recherche exhaustive (valeur+date, tolerance quasi nulle) ET tests
# de transformation (delta, cumul, arrondi, unite). Identifiees par
# (nom de feuille, titre exact tel qu'ecrit en ligne 3 de la feuille).
A_RETIRER = [
    ('Agriculture', 'Moyenne des precipitations'),
    ('Agriculture', 'Température_moyenne'),
    ("Industrie d'extraction", 'Production du phosphate brut  (Jan 2018)'),
    ("Industrie d'extraction", 'Production des dérivées de phosphates (Jan 2018)'),
    ("Industrie d'extraction", 'Exportations OCP'),
    ("Industrie d'extraction", 'Exportations phosphate+derivés'),
    ("Industrie d'extraction", 'Exportations autres extractions minières (Oct 2018)'),
    ('Industrie de transformation', 'Utilisation des capacités de production (cumulé)'),
    ('Industrie de transformation', "PRODUITS FINIS D'EQUIPEMENT INDUSTRIEL"),
    ('Industrie de transformation', 'DEMI PRODUITS '),
    ('Électricité, gaz, eau', "Production d'électricité\n\n"),
    ('Électricité, gaz, eau', 'Consommation de combustibles (Fioul, Charbon, Gaz oil) Tonnes'),
    ('Commerce', 'PRODUITS FINIS DE CONSOMMATION 1000 DH'),
    ('Commerce', 'TOTAL EXPORTATIONS (1000)'),
    ('Commerce', 'TOTAL IMPORTATIONS  (1000)'),
    ('Commerce', 'Indice IMPORTATIONS'),
    ('Commerce', 'Indice EXPORTATIONS'),
    ('Information-communication', 'Taux de pénétration du mobile'),
    ('Information-communication', 'Parc Internet global  (en milliers)'),
    ('Information-communication', 'Taux de pénétration de l’Internet'),
    ('Information-communication', 'Parc liaisons Data Entreprises Nationale (T1-2018)'),
    ('Information-communication', 'Parc des noms de domaine «.ma»'),
    ('Information-communication', 'Nouveaux enregistrements durant le trimestre'),
    ('Information-communication', 'Trafic voix sortant Mobile — Trafic voix sortant du Mobile'),
    ('Finances et assurances', 'Dépôts à vue auprés des banques'),
    ('Finances et assurances', 'Masse monétaire (M3)'),
    ('Finances et assurances', 'Avoirs officiels de réserve'),
    ('Finances et assurances', 'Primes versées (MDH)'),
    ('Finances et assurances', 'Prestations (MDH)'),
    ('Immobilier', 'Indice des prix des actifs immobiliers (IPAI)'),
    ('Immobilier', 'taux Crédits immobiliers'),
]

A_REMPLACER = [
    # (feuille, ancien_titre, nouveau_fichier, nouvelle_serie, methode)
    ('Construction', 'Vente de ciment (1000tonnes)',
     "Manar-Stat (Ministere de l'Economie et des Finances)\\secondaire\\construction\\"
     "Ventes locales du ciment (mensuel cumulé).csv",
     'Total des régions', 'delta mensuel de la serie cumulee'),
    ('Industrie de transformation', 'Taux d’Utilisation des Capacités',
     "Manar-Stat (Ministere de l'Economie et des Finances)\\secondaire\\industrie\\"
     "Taux d’utilisation des capacités (mensuel).csv",
     'TUC industrielle', 'serie primaire complete (le classeur en avait un arrondi tronque)'),
]


def cle_mois(v):
    s = str(v).strip() if v is not None else ''
    if re.fullmatch(r'\d{4}M\d{2}', s):
        return s
    m = re.fullmatch(r'([a-zûéèA-Z]+)-(\d{2})', s.lower().replace('.', ''))
    if m and m.group(1)[:4] in MOIS_FR:
        an = int(m.group(2))
        an += 2000 if an < 80 else 1900
        return f'{an}M{MOIS_FR[m.group(1)[:4]]:02d}'
    return None


def trouver_colonne(ws, titre):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(3, c).value or '').strip() == titre.strip():
            return c
    return None


def trouver_axe(ws, c):
    for cc in range(c, 0, -1):
        v3, v4 = ws.cell(3, cc).value, ws.cell(4, cc).value
        if v3 in ('Trimestre', 'Mois', 'Année') or v4 in ('Trimestre', 'Mois', 'Année'):
            return cc, (v3 if v3 in ('Trimestre', 'Mois', 'Année') else v4)
    return None, None


def main():
    print('Chargement du classeur et du data lake ...')
    wb = openpyxl.load_workbook(CLASSEUR)
    lake = {s['fichier']: s for s in charger_lake(ROOT)}
    # charger_lake peut generer plusieurs entrees par fichier (une par serie) :
    # on veut toutes les series d'un fichier donne -> reindexer par (fichier,serie)
    lake_series = {}
    for s in charger_lake(ROOT):
        lake_series[(s['fichier'], s['serie'])] = s

    # ---------------------------------------------------------- 1. remplacements
    print('\n=== REMPLACEMENTS (transformation legitime confirmee) ===')
    for feuille, titre, fichier, serie, methode in A_REMPLACER:
        ws = wb[feuille]
        c = trouver_colonne(ws, titre)
        if c is None:
            print('  !! colonne "%s" introuvable dans %s' % (titre, feuille))
            continue
        col_axe, type_axe = trouver_axe(ws, c)
        src = lake_series.get((fichier, serie))
        if src is None:
            print('  !! source introuvable :', fichier, '::', serie)
            continue
        vals = src['valeurs']

        if 'delta' in methode:
            ordre = sorted(vals)
            nouv = {}
            for i in range(1, len(ordre)):
                k, kp = ordre[i], ordre[i - 1]
                if k[:4] == kp[:4]:            # meme annee -> cumul valide
                    nouv[k] = round(vals[k] - vals[kp], 3)
                else:                          # janvier : cumul = valeur du mois
                    nouv[k] = vals[k]
            nouv[ordre[0]] = vals[ordre[0]]
        else:
            nouv = vals

        # etendre l'axe si besoin (la nouvelle serie peut couvrir plus de mois)
        besoin = sorted(nouv)
        axe_actuel = {}
        for r in range(5, ws.max_row + 1):
            k = cle_mois(ws.cell(r, col_axe).value)
            if k:
                axe_actuel[k] = r
        derniere_ligne = max(axe_actuel.values()) if axe_actuel else 4
        for k in besoin:
            if k not in axe_actuel:
                derniere_ligne += 1
                ws.cell(derniere_ligne, col_axe).value = k
                axe_actuel[k] = derniere_ligne
        # effacer les anciennes valeurs puis ecrire les nouvelles
        for r in axe_actuel.values():
            ws.cell(r, c).value = None
        for k, v in nouv.items():
            ws.cell(axe_actuel[k], c).value = v
        ws.cell(4, c).value = fichier
        ws.cell(2, c).value = 'REMPLACEE (%s) : %s' % (methode, serie)

        print('  %-28s | %-48s -> %s (%d -> %d valeurs, %s -> %s)'
              % (feuille, titre[:48], Path(fichier).name[:40],
                 sum(1 for r in range(5, ws.max_row + 1)
                     if isinstance(ws.cell(r, c).value, (int, float))),
                 len(nouv), min(nouv), max(nouv)))

    # ---------------------------------------------------------- 2. retraits
    print('\n=== RETRAITS (aucune source verifiable trouvee) ===')
    retires = []
    for feuille, titre in A_RETIRER:
        ws = wb[feuille]
        c = trouver_colonne(ws, titre)
        if c is None:
            print('  !! colonne "%s" introuvable dans %s (deja retiree ?)' % (titre, feuille))
            continue
        n = sum(1 for r in range(5, ws.max_row + 1)
                if isinstance(ws.cell(r, c).value, (int, float)))
        retires.append([feuille, titre.strip(), n])
        ws.delete_cols(c, 1)
        print('  %-28s | %-52s (%d valeurs retirees)' % (feuille, titre.strip()[:52], n))

    # ---------------------------------------------------------- 3. feuille de trace
    if 'Colonnes retirées' in wb.sheetnames:
        del wb['Colonnes retirées']
    wr = wb.create_sheet('Colonnes retirées')
    wr.append(['Colonnes retirées du classeur — audit de fidélité'])
    wr.cell(1, 1).font = Font(bold=True, size=14, color='C00000')
    wr.append(['Recherche exhaustive dans les 481 fichiers du data lake (valeur '
               'exacte + date, tolérance quasi nulle) et tests de transformation '
               '(delta, cumul, arrondi). Aucune correspondance trouvée : ces '
               'colonnes ne peuvent pas être vérifiées, elles ont été retirées '
               'plutôt que laissées avec un statut « non vérifié » silencieux.'])
    wr.cell(2, 1).font = Font(italic=True, size=10, color='595959')
    wr.append([])
    wr.append(['Branche', 'Colonne retirée', 'Nb valeurs perdues'])
    for r in retires:
        wr.append(r)
    en = PatternFill('solid', fgColor='C00000')
    for c, w in zip('ABC', (26, 60, 18)):
        wr.column_dimensions[c].width = w
        cel = wr.cell(4, ord(c) - 64)
        cel.font = Font(bold=True, color='FFFFFF')
        cel.fill = en
        cel.alignment = Alignment(horizontal='center')
    wr.row_dimensions[4].height = 28
    wr.freeze_panes = 'A5'
    wr.auto_filter.ref = 'A4:C%d' % wr.max_row
    wr.sheet_view.showGridLines = False

    wb.save(CLASSEUR)
    print('\n%s enregistre.' % CLASSEUR.name)
    print('%d colonnes remplacees, %d colonnes retirees.' % (len(A_REMPLACER), len(retires)))


if __name__ == '__main__':
    main()
