# -*- coding: utf-8 -*-
"""
Reconstruit Etude_sectorielle_Maroc_1.xlsx en realignant chaque colonne sur les
dates reelles de sa source.

Probleme corrige
----------------
Dans le classeur d'origine, les series ont ete collees a partir de la premiere
ligne de leur bloc au lieu d'etre jointes sur la date. Toute serie dont la
premiere periode disponible est posterieure a l'origine de l'axe se retrouve
donc decalee (ex. l'IPI base 2015, qui commence en 2016T1, apparaissait a
partir de T4-2006 : un decalage de 37 trimestres).

Methode
-------
1. Catalogue de toutes les series sources (Manar-Stat via manar_panel.csv,
   Bank Al-Maghrib, IPAI).
2. Pour chaque colonne du classeur, identification de la serie source par
   comparaison des VALEURS (les valeurs sont exactes, seules les dates sont
   fausses) : on teste l'appariement par date et l'appariement par position.
3. Reecriture de la colonne par jointure sur la date.
4. Feuille "Audit" : provenance et statut de chaque colonne.

Sortie : Etude_sectorielle_Maroc_2.xlsx
"""

import csv
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
SRC_XLSX = ROOT / 'Etude_sectorielle_Maroc_1.xlsx'
OUT_XLSX = ROOT / 'Etude_sectorielle_Maroc_2.xlsx'
PANEL = ROOT / 'Bases consolidees (multi-sources)' / 'manar_panel.csv'
BAM = ROOT / 'Bank Al-Maghrib et ANCFCC' / 'credit_bancaire_par_branche'
IPAI = ROOT / 'Bank Al-Maghrib et ANCFCC' / 'immobilier_ipai' / 'ipai_variations_propre.csv'

MOIS_FR = {'janv': 1, 'févr': 2, 'fevr': 2, 'mars': 3, 'avr': 4, 'mai': 5,
           'juin': 6, 'juil': 7, 'août': 8, 'aout': 8, 'sept': 9, 'oct': 10,
           'nov': 11, 'déc': 12, 'dec': 12}


# --------------------------------------------------------------- periodes
def cle_trim(v):
    if isinstance(v, str):
        m = re.fullmatch(r'\s*T([1-4])-(\d{4})\s*', v)
        if m:
            return f'{m.group(2)}T{m.group(1)}'
    return None


def cle_mois(v):
    if isinstance(v, datetime):
        return f'{v.year}M{v.month:02d}'
    if isinstance(v, str):
        s = v.strip().lower().replace('.', '')
        m = re.fullmatch(r'([a-zûéè]+)-(\d{2})', s)
        if m and m.group(1) in MOIS_FR:
            a = int(m.group(2))
            a += 2000 if a < 80 else 1900
            return f'{a}M{MOIS_FR[m.group(1)]:02d}'
        m = re.fullmatch(r'(\d{4})M(\d{2})', v.strip())
        if m:
            return v.strip()
    return None


def _norm(s):
    import unicodedata
    s = unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode().lower()
    return re.sub(r'[^a-z0-9]', '', s)


def cle_axe(v, type_axe):
    return cle_trim(v) if type_axe == 'Trimestre' else cle_mois(v)


def trim_depuis_date(txt):
    m = re.match(r'(\d{4})-(\d{2})-\d{2}', str(txt))
    if not m:
        return None
    a, mo = int(m.group(1)), int(m.group(2))
    return f'{a}T{(mo - 1) // 3 + 1}'


def mois_depuis_date(txt):
    m = re.match(r'(\d{4})-(\d{2})-\d{2}', str(txt))
    return f'{int(m.group(1))}M{int(m.group(2)):02d}' if m else None


# --------------------------------------------------------------- catalogue
def nombre(t):
    t = str(t).strip().replace(' ', '').replace('\xa0', '')
    if not t or t.lower() in ('null', 'nan', '-', '..'):
        return None
    t = t.replace(',', '.')
    try:
        return float(t)
    except ValueError:
        return None


def charger_catalogue():
    """-> liste de dict {origine, tableau, serie, freq, valeurs{periode: val}}"""
    cat = []

    # 1. Manar-Stat, depuis le panel harmonise (deja date-normalise).
    #    La cle est recalculee depuis date_fin : les series cumulees portent des
    #    libelles ("Janv:Juin-2016") qui ne se joignent ni ne se trient.
    if PANEL.exists():
        agg = defaultdict(dict)
        meta = {}
        with open(PANEL, encoding='utf-8-sig', newline='') as fh:
            for row in csv.DictReader(fh, delimiter=';'):
                cle = (row['tableau'], row['serie'])
                v = nombre(row['valeur'])
                if v is None:
                    continue
                f = row['frequence']
                fin = row['date_fin']          # AAAA-MM-JJ
                an, mo = int(fin[:4]), int(fin[5:7])
                if 'rimestr' in f:
                    k = f'{an}T{(mo - 1) // 3 + 1}'
                elif 'ensuel' in f:
                    k = f'{an}M{mo:02d}'
                else:
                    continue                    # annuel / campagne : hors axes du classeur
                agg[cle][k] = v
                meta[cle] = f
        for (tab, ser), vals in agg.items():
            cat.append({'origine': 'Manar-Stat', 'tableau': tab, 'serie': ser,
                        'freq': meta[(tab, ser)], 'valeurs': vals})

    # 2. Bank Al-Maghrib
    if BAM.is_dir():
        for f in sorted(BAM.iterdir()):
            if f.suffix.lower() != '.csv':
                continue
            rows = list(csv.reader(open(f, encoding='utf-8-sig'), delimiter=','))
            ligne_dates = None
            for i, r in enumerate(rows[:6]):
                if sum(1 for c in r if re.match(r'\d{4}-\d{2}-\d{2}', str(c))) > 3:
                    ligne_dates = i
                    break
            if ligne_dates is None:
                continue
            entete = rows[ligne_dates]
            trimestriel = 'trimestriel' in f.name.lower() or 'Ventilation des' in f.name
            cles = []
            for c in entete:
                k = trim_depuis_date(c) if trimestriel else mois_depuis_date(c)
                cles.append(k)
            # si les dates ne sont pas espacees de 3 mois -> mensuel
            for r in rows[ligne_dates + 1:]:
                if not r or not r[0].strip():
                    continue
                vals = {}
                for j, k in enumerate(cles):
                    if k and j < len(r):
                        v = nombre(r[j])
                        if v is not None:
                            vals[k] = v
                if len(vals) >= 8:
                    cat.append({'origine': 'Bank Al-Maghrib', 'tableau': f.name,
                                'serie': r[0].strip(),
                                'freq': 'Trimestrielle' if trimestriel else 'Mensuelle',
                                'valeurs': vals})

    # 2 bis. Balayage generique de TOUT le data lake : les series qui
    #        n'etaient pas dans ma collecte (Office des Changes, ONEE, OCP,
    #        tourisme, ANRT, ciment, precipitations...) vivent dans des CSV
    #        exportes tels quels, hors de manar_panel.csv.
    try:
        from catalogue_lake import charger_lake
        for s in charger_lake(ROOT):
            cat.append({'origine': 'data lake', 'tableau': s['fichier'],
                        'serie': s['serie'], 'freq': s['freq'],
                        'valeurs': s['valeurs']})
    except Exception as e:                                   # pragma: no cover
        print('  (balayage du data lake indisponible : %s)' % e)

    # 3. IPAI (extrait des bulletins Bank Al-Maghrib / ANCFCC)
    if IPAI.exists():
        agg = defaultdict(dict)
        with open(IPAI, encoding='utf-8-sig', newline='') as fh:
            for row in csv.DictReader(fh, delimiter=';'):
                k = cle_trim(row.get('trimestre', ''))
                if not k:
                    continue
                cat_, ind = row.get('categorie', ''), row.get('indicateur', '')
                for champ, suffixe in (('variation_trimestrielle', 'var. trimestrielle'),
                                       ('variation_annuelle', 'var. annuelle')):
                    v = nombre(row.get(champ, ''))
                    if v is not None:
                        agg[(f'{cat_} — {ind}', suffixe)][k] = v
        for (serie, suffixe), vals in agg.items():
            if len(vals) >= 5:
                cat.append({'origine': 'Bank Al-Maghrib / ANCFCC (IPAI)',
                            'tableau': IPAI.name, 'serie': f'{serie} [{suffixe}]',
                            'freq': 'Trimestrielle', 'valeurs': vals})
    return cat


# --------------------------------------------------------------- appariement
def apparier(observees, cat, type_axe):
    """observees : [(periode_axe, valeur)] telles qu'ecrites dans le classeur.
    Renvoie (meilleur, methode, score) ou (None, None, 0)."""
    if len(observees) < 5:
        return None, None, 0
    ampl = sum(abs(v) for _, v in observees[:8]) / min(8, len(observees)) or 1.0
    tol = max(0.051, ampl * 0.002)
    attendu = 'Trimestrielle' if type_axe == 'Trimestre' else 'Mensuelle'

    best = (None, None, 0)
    for s in cat:
        if attendu == 'Trimestrielle' and 'rimestr' not in s['freq']:
            continue
        if attendu == 'Mensuelle' and 'ensuel' not in s['freq']:
            continue
        vals = s['valeurs']
        ordre = sorted(vals)  # les cles YYYYTn / YYYYMmm se trient bien
        d = sum(1 for p, v in observees if p in vals and abs(v - vals[p]) <= tol)
        n = min(len(observees), len(ordre))
        q = sum(1 for i in range(n) if abs(observees[i][1] - vals[ordre[i]]) <= tol)
        if d >= q and d > best[2]:
            best = (s, 'date', d)
        elif q > d and q > best[2]:
            best = (s, 'position', q)
    # Seuil volontairement exigeant : reecrire une colonne sur un appariement
    # mediocre reviendrait a remplacer des donnees legitimes par une serie
    # peut-etre differente. En dessous, on prefere ne pas toucher et signaler.
    seuil = max(8, int(0.85 * len(observees)))
    return best if best[2] >= seuil else (None, None, best[2])


# --------------------------------------------------------------- principal
def main():
    if not SRC_XLSX.exists():
        sys.exit('classeur source introuvable : %s' % SRC_XLSX)
    print('Chargement du catalogue de sources ...')
    cat = charger_catalogue()
    from collections import Counter as _C
    print('  %d series sources : %s' % (len(cat), dict(_C(s['origine'] for s in cat))))

    wb = openpyxl.load_workbook(SRC_XLSX)
    audit = []

    for feuille in wb.sheetnames:
        if feuille in ('Sommaire', 'Métadonnées', 'Audit'):
            continue
        ws = wb[feuille]
        # reperage des blocs : colonne d'axe -> colonnes de donnees a sa droite
        axes = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(4, c).value
            if isinstance(v, str) and v.strip() in ('Trimestre', 'Mois'):
                axes.append((c, v.strip()))
        for idx, (col_axe, type_axe) in enumerate(axes):
            fin = axes[idx + 1][0] - 1 if idx + 1 < len(axes) else ws.max_column
            lignes = []
            for r in range(5, ws.max_row + 1):
                k = cle_axe(ws.cell(r, col_axe).value, type_axe)
                if k:
                    lignes.append((r, k))
            if not lignes:
                continue
            for c in range(col_axe + 1, fin + 1):
                nom = ws.cell(3, c).value or ws.cell(4, c).value
                obs = []
                for r, k in lignes:
                    v = ws.cell(r, c).value
                    if isinstance(v, (int, float)):
                        obs.append((k, float(v)))
                if not obs:
                    continue
                if c == 2 and col_axe == 1:      # colonne VA cible : pas de source externe
                    audit.append([feuille, str(nom)[:80], 'VA cible (HCP)', '', 'conservee',
                                  len(obs), obs[0][0], obs[-1][0], ''])
                    continue
                s, methode, score = apparier(obs, cat, type_axe)
                if s is None:
                    audit.append([feuille, str(nom)[:80], '?', '', 'NON IDENTIFIEE — conservee',
                                  len(obs), obs[0][0], obs[-1][0],
                                  'meilleur score %d / %d' % (score, len(obs))])
                    continue
                vals = s['valeurs']
                # reecriture par jointure sur la date
                ecrits = 0
                for r, k in lignes:
                    ws.cell(r, c).value = vals.get(k)
                    if k in vals:
                        ecrits += 1
                dispo = sorted(vals)
                audit.append([feuille, str(nom)[:80],
                              '%s — %s' % (s['origine'], s['tableau'][:60]),
                              s['serie'][:60],
                              'REALIGNEE' if methode == 'position' else 'verifiee (deja alignee)',
                              ecrits, dispo[0], dispo[-1], ''])
        print('  %-30s %d colonnes traitees' % (feuille, sum(1 for a in audit if a[0] == feuille)))

    # --- mise a jour de la feuille Metadonnees -------------------------------
    # Elle annoncait l'origine de l'axe comme date de debut ("T4-2006",
    # "dec-07") alors que les series commencent souvent bien plus tard.
    reels = {}          # (feuille, libelle normalise) -> (debut, fin, n)
    for feuille in wb.sheetnames:
        if feuille in ('Sommaire', 'Métadonnées', 'Audit'):
            continue
        ws = wb[feuille]
        axes = [(c, ws.cell(4, c).value.strip())
                for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(4, c).value, str)
                and ws.cell(4, c).value.strip() in ('Trimestre', 'Mois')]
        for idx, (col_axe, type_axe) in enumerate(axes):
            fin_bloc = axes[idx + 1][0] - 1 if idx + 1 < len(axes) else ws.max_column
            lignes = [(r, cle_axe(ws.cell(r, col_axe).value, type_axe))
                      for r in range(5, ws.max_row + 1)]
            lignes = [(r, k) for r, k in lignes if k]
            for c in range(col_axe + 1, fin_bloc + 1):
                nom = ws.cell(3, c).value
                if not nom:
                    continue
                per = [k for r, k in lignes
                       if isinstance(ws.cell(r, c).value, (int, float))]
                if per:
                    reels[(feuille, _norm(nom))] = (per[0], per[-1], len(per))

    wm = wb['Métadonnées'] if 'Métadonnées' in wb.sheetnames else None
    maj = rate = 0
    if wm is not None:
        for r in range(2, wm.max_row + 1):
            branche = wm.cell(r, 1).value
            serie = wm.cell(r, 2).value
            if not branche or not serie:
                continue
            cible = None
            for (f, n), v in reels.items():
                if f == branche and (n == _norm(serie)
                                     or n.startswith(_norm(serie)[:20])
                                     or _norm(serie).startswith(n[:20])):
                    cible = v
                    break
            if cible:
                wm.cell(r, 5).value, wm.cell(r, 6).value, wm.cell(r, 7).value = cible
                maj += 1
            else:
                rate += 1
        print('\nMetadonnees : %d lignes mises a jour, %d non appariees' % (maj, rate))

    # feuille d'audit
    if 'Audit' in wb.sheetnames:
        del wb['Audit']
    wa = wb.create_sheet('Audit')
    entete = ['Feuille', 'Colonne', 'Fichier source', 'Serie source', 'Statut',
              'Nb valeurs ecrites', 'Debut source', 'Fin source', 'Remarque']
    wa.append(entete)
    for c in range(1, len(entete) + 1):
        wa.cell(1, c).font = Font(bold=True)
    rouge = PatternFill('solid', fgColor='FFC7CE')
    orange = PatternFill('solid', fgColor='FFEB9C')
    for ligne in audit:
        wa.append(ligne)
        r = wa.max_row
        if ligne[4].startswith('REALIGNEE'):
            wa.cell(r, 5).fill = orange
        elif ligne[4].startswith('NON IDENTIFIEE'):
            wa.cell(r, 5).fill = rouge
    for c, l in zip('ABCDEFGHI', (22, 52, 46, 34, 26, 12, 12, 12, 26)):
        wa.column_dimensions[c].width = l

    wb.save(OUT_XLSX)

    n_re = sum(1 for a in audit if a[4].startswith('REALIGNEE'))
    n_ok = sum(1 for a in audit if a[4].startswith('verifiee'))
    n_no = sum(1 for a in audit if a[4].startswith('NON'))
    n_va = sum(1 for a in audit if a[4] == 'conservee')
    print('\n%s' % OUT_XLSX.name)
    print('  colonnes realignees        : %d' % n_re)
    print('  colonnes deja correctes    : %d' % n_ok)
    print('  colonnes non identifiees   : %d' % n_no)
    print('  colonnes VA cible (gardees): %d' % n_va)


if __name__ == '__main__':
    main()
