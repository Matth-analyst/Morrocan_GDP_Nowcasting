# -*- coding: utf-8 -*-
"""
Restaure les 20 dernieres colonnes retirees (hors IPAI, traite a part).

Contexte : elles proviennent de BDD SECTORIEL_MENSUEL/TRIM.csv, elles-memes
converties depuis "BDD SECTORIEL.xlsx", un fichier construit par
Alim ATCHADAM (cree le 22/07/2026, modifie le 03/08/2026, transmis par
WhatsApp). Inspection de ce fichier original : 2 feuilles (MENSUEL, TRIM),
aucun commentaire de cellule, aucune formule -- valeurs saisies directement,
sans metadonnee de source supplementaire au-dela de ce que les CSV
contenaient deja.

L'utilisateur a decide, en connaissance de cause, de reintegrer ces series
malgre l'absence de fichier primaire permettant de verifier l'alignement des
dates. Elles sont restaurees a l'identique de la version du classeur juste
avant leur retrait, avec un statut explicite.
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'
AVANT = Path(r'C:\Users\HP\AppData\Local\Temp\claude_restore\avant_retrait.xlsx')

AUTEUR = 'Alim ATCHADAM (fichier BDD SECTORIEL.xlsx, créé 22/07/2026)'
STATUT = 'confirmé par l’auteur — date non vérifiée'
ORANGE = PatternFill('solid', fgColor='FFE699')

A_RESTAURER = [
    ("Industrie d'extraction", 'Production du phosphate brut  (Jan 2018)', 'OCP / Manar-Stat'),
    ("Industrie d'extraction", 'Production des dérivées de phosphates (Jan 2018)', 'OCP / Manar-Stat'),
    ("Industrie d'extraction", 'Exportations OCP', 'OCP / Manar-Stat'),
    ("Industrie d'extraction", 'Exportations phosphate+derivés', 'OCP / Manar-Stat'),
    ("Industrie d'extraction", 'Exportations autres extractions minières (Oct 2018)', 'OCP / Manar-Stat'),
    ('Industrie de transformation', 'Utilisation des capacités de production (cumulé)', 'Manar-Stat'),
    ('Industrie de transformation', "PRODUITS FINIS D'EQUIPEMENT INDUSTRIEL", 'Office des Changes / Manar-Stat'),
    ('Industrie de transformation', 'DEMI PRODUITS ', 'Office des Changes / Manar-Stat'),
    ('Électricité, gaz, eau', "Production d'électricité\n\n", 'ONEE / Manar-Stat'),
    ('Électricité, gaz, eau', 'Consommation de combustibles (Fioul, Charbon, Gaz oil) Tonnes', 'ONEE / Manar-Stat'),
    ('Commerce', 'PRODUITS FINIS DE CONSOMMATION 1000 DH', 'Office des Changes / Manar-Stat'),
    ('Commerce', 'TOTAL EXPORTATIONS (1000)', 'Office des Changes / Manar-Stat'),
    ('Commerce', 'TOTAL IMPORTATIONS  (1000)', 'Office des Changes / Manar-Stat'),
    ('Information-communication', 'Taux de pénétration du mobile', 'ANRT / Manar-Stat'),
    ('Information-communication', 'Parc Internet global  (en milliers)', 'ANRT / Manar-Stat'),
    ('Information-communication', 'Taux de pénétration de l’Internet', 'ANRT / Manar-Stat'),
    ('Information-communication', 'Parc liaisons Data Entreprises Nationale (T1-2018)', 'ANRT / Manar-Stat'),
    ('Information-communication', 'Parc des noms de domaine «.ma»', 'ANRT / Manar-Stat'),
    ('Information-communication', 'Nouveaux enregistrements durant le trimestre', 'ANRT / Manar-Stat'),
    ('Information-communication', 'Trafic voix sortant Mobile — Trafic voix sortant du Mobile', 'ANRT / Manar-Stat'),
]


def trouver_colonne(ws, titre):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(3, c).value or '').strip() == titre.strip():
            return c
    return None


def trouver_axe(ws, c):
    for cc in range(c, 0, -1):
        v3, v4 = ws.cell(3, cc).value, ws.cell(4, cc).value
        if v3 in ('Trimestre', 'Mois', 'Année') or v4 in ('Trimestre', 'Mois', 'Année'):
            return cc, (v3 if v3 in ('Trimestre', 'Mois', 'Année') else v4)
    return None, None


def main():
    wb_avant = openpyxl.load_workbook(AVANT, data_only=True)
    wb = openpyxl.load_workbook(CLASSEUR)

    restaurees = []
    for feuille, titre, inst in A_RESTAURER:
        ws_av = wb_avant[feuille]
        c_av = trouver_colonne(ws_av, titre)
        if c_av is None:
            print('  !! introuvable (version avant retrait) :', feuille, '|', titre)
            continue
        axe_av, type_axe = trouver_axe(ws_av, c_av)
        paires = []
        for r in range(5, ws_av.max_row + 1):
            k = ws_av.cell(r, axe_av).value
            v = ws_av.cell(r, c_av).value
            if k is not None and isinstance(v, (int, float)):
                paires.append((k, v))
        if not paires:
            print('  !! aucune valeur :', feuille, '|', titre)
            continue

        ws = wb[feuille]
        axe_actuel = None
        for c in range(1, ws.max_column + 1):
            v3, v4 = ws.cell(3, c).value, ws.cell(4, c).value
            if v3 == type_axe or v4 == type_axe:
                axe_actuel = c
        if axe_actuel is None:
            print('  !! aucun axe %s dans %s' % (type_axe, feuille))
            continue
        c_ins = axe_actuel + 1
        while not all(ws.cell(r, c_ins).value in (None, '') for r in range(1, 12)):
            c_ins += 1
        ws.insert_cols(c_ins, 1)

        ws.cell(3, c_ins).value = titre
        ws.cell(4, c_ins).value = 'BDD SECTORIEL (source institutionnelle confirmée : %s)' % inst
        ws.cell(2, c_ins).value = 'RESTAUREE — %s — auteur : %s' % (STATUT, AUTEUR)
        ws.cell(3, c_ins).fill = ORANGE

        lignes_axe = {}
        for r in range(5, ws.max_row + 1):
            k = ws.cell(r, axe_actuel).value
            if k is not None:
                lignes_axe[k] = r
        derniere = max(lignes_axe.values()) if lignes_axe else 4
        for k, v in paires:
            if k not in lignes_axe:
                derniere += 1
                ws.cell(derniere, axe_actuel).value = k
                lignes_axe[k] = derniere
            ws.cell(lignes_axe[k], c_ins).value = v

        restaurees.append([feuille, titre.strip(), inst, len(paires)])
        print('  %-28s | %-52s | %-30s | %d valeurs'
              % (feuille, titre.strip()[:52], inst[:30], len(paires)))

    wb.save(CLASSEUR)
    print('\n%d colonnes restaurees, enregistrees dans %s' % (len(restaurees), CLASSEUR.name))
    return restaurees


if __name__ == '__main__':
    main()
