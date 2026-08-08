# -*- coding: utf-8 -*-
"""
Complete Etude_sectorielle_Maroc_2.xlsx avec les series du data lake qui
peuvent rejoindre celles deja retenues.

Principe : le contenu existant de _2 n'est PAS touche. Les series nouvelles
sont ajoutees a droite de chaque feuille, dans deux blocs "AJOUTS" munis de
leur propre axe (trimestriel et mensuel), avec leur source.

Sortie : Etude_sectorielle_Maroc_2_complete.xlsx
"""

import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parent))
from catalogue_lake import charger_lake                        # noqa: E402
from enrichir_classeur import classer, sansacc, BRANCHES, periodes_trim, periodes_mois  # noqa: E402

SOURCE = ROOT / 'Etude_sectorielle_Maroc_2.xlsx'
SORTIE = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'
IPAI = ROOT / 'Bank Al-Maghrib et ANCFCC' / 'immobilier_ipai' / 'ipai_variations_propre.csv'
VA_FICHIER = 'PIB trimestriel_PIB base2014 rétropolé 28 branc.csv'

EXCLURE_FICHIER = ['bdd sectoriel', 'etude_sectorielle', 'manar_panel',
                   '_transfert', 'graphify-out']
EXCLURE_SERIE = [r'^v\s*(\(\d+\))?$', r'^vl\s*(\(\d+\))?$', r'^unnamed',
                 r'^colonne\d*$', r'^[a-z]$']

# styles
F_TITRE = Font(bold=True, size=12, color='7030A0')
F_BLOC = Font(bold=True, size=11, color='FFFFFF')
F_COL = Font(bold=True, size=10)
F_SRC = Font(italic=True, size=8, color='7F7F7F')
F_AXE = Font(bold=True, size=10, color='7030A0')
BL_T = PatternFill('solid', fgColor='7030A0')
BL_M = PatternFill('solid', fgColor='C55A11')
BL_C = PatternFill('solid', fgColor='BF8F00')      # cibles alternatives
BL_A = PatternFill('solid', fgColor='548235')      # bloc annuel
H_T = PatternFill('solid', fgColor='E4DFEC')
H_M = PatternFill('solid', fgColor='FBE5D6')
H_C = PatternFill('solid', fgColor='FFF2CC')
H_A = PatternFill('solid', fgColor='E2EFDA')
AXE_F = PatternFill('solid', fgColor='F2F2F2')
CENTRE = Alignment(horizontal='center', vertical='center', wrap_text=True)
HAUT = Alignment(horizontal='center', vertical='top', wrap_text=True)
BAS = Border(bottom=Side(style='thin', color='BFBFBF'))


def lire_ipai():
    """Le fichier IPAI est en format long : une ligne par trimestre x categorie
    x indicateur. Le lecteur generique (format large) ne le voit pas."""
    if not IPAI.exists():
        return []
    agg = defaultdict(dict)
    with open(IPAI, encoding='utf-8-sig', newline='') as fh:
        for row in csv.DictReader(fh, delimiter=';'):
            m = re.fullmatch(r'T([1-4])-(\d{4})', str(row.get('trimestre', '')).strip())
            if not m:
                continue
            k = f'{m.group(2)}T{m.group(1)}'
            cat = (row.get('categorie') or '').strip()
            ind = (row.get('indicateur') or '').strip()
            for champ, suff in (('variation_trimestrielle', 'var. trim.'),
                                ('variation_annuelle', 'var. annuelle')):
                t = (row.get(champ) or '').strip().replace(',', '.')
                if not t:
                    continue
                try:
                    v = float(t)
                except ValueError:
                    continue
                agg[(f'IPAI {cat} — {ind} [{suff}]',)][k] = v
    out = []
    for (nom,), vals in agg.items():
        if len(vals) >= 8:
            out.append({'fichier': str(IPAI.relative_to(ROOT)), 'serie': nom,
                        'freq': 'Trimestrielle', 'valeurs': vals})
    return out


PANEL = ROOT / 'Bases consolidees (multi-sources)' / 'manar_panel.csv'
MAX_ANNUEL = 50          # plafond de series annuelles par branche


def lire_panel_annuel():
    """Series annuelles et de campagne agricole, indexees par ANNEE.

    Une campagne 2020/2021 est placee sur son annee de recolte (2021), ce qui
    permet de la juxtaposer aux series annuelles sur un meme axe."""
    if not PANEL.exists():
        return []
    agg = defaultdict(dict)
    meta = {}
    with open(PANEL, encoding='utf-8-sig', newline='') as fh:
        for row in csv.DictReader(fh, delimiter=';'):
            if row['frequence'] not in ('Annuelle', 'Année à cheval'):
                continue
            try:
                v = float(row['valeur'])
                an = int(row['annee'])
            except (ValueError, KeyError):
                continue
            cle = (row['tableau'], row['serie'])
            agg[cle][an] = v
            meta[cle] = (row['frequence'], row['rubrique'])
    out = []
    for (tab, ser), vals in agg.items():
        freq, rub = meta[(tab, ser)]
        if len(vals) >= 12 and max(vals) >= 2015:
            out.append({'fichier': '%s (manar_panel)' % tab, 'serie': ser,
                        'freq': freq, 'rubrique': rub, 'valeurs': vals})
    return out


def variante_pib(nom_fichier):
    """Etiquette courte pour distinguer les variantes de PIB."""
    n = nom_fichier.replace('PIB trimestriel', '').replace('.csv', '')
    n = n.strip(' _-')
    return n or 'niveau base 2014'


def deja_utilisees(wb2):
    """(nom de fichier, serie) deja mobilises par _2, d'apres sa feuille Audit."""
    vus = set()
    if 'Audit' not in wb2.sheetnames:
        return vus
    for r in wb2['Audit'].iter_rows(min_row=2, values_only=True):
        if r and r[2] and r[3]:
            vus.add((Path(str(r[2]).split('—')[-1].strip()).name, str(r[3]).strip()))
            vus.add(str(r[3]).strip())
    return vus


def ecrire_bloc(ws, col0, titre, fill_bandeau, fill_col, axe_nom, periodes, series):
    n = len(series)
    if n:
        ws.merge_cells(start_row=1, start_column=col0, end_row=1, end_column=col0 + n)
    c = ws.cell(1, col0)
    c.value = titre
    c.font = F_BLOC
    c.fill = fill_bandeau
    c.alignment = CENTRE

    ws.cell(3, col0).value = axe_nom
    ws.cell(3, col0).font = F_AXE
    ws.cell(3, col0).alignment = CENTRE
    ws.cell(3, col0).fill = AXE_F
    ws.column_dimensions[get_column_letter(col0)].width = 11
    for i, p in enumerate(periodes):
        cel = ws.cell(5 + i, col0)
        cel.value = p
        cel.font = F_AXE
        cel.fill = AXE_F
        cel.alignment = Alignment(horizontal='center')

    for j, s in enumerate(series, start=1):
        cc = col0 + j
        ws.column_dimensions[get_column_letter(cc)].width = 15
        t = ws.cell(3, cc)
        t.value = s['serie']
        t.font = F_COL
        t.fill = fill_col
        t.alignment = HAUT
        f = ws.cell(4, cc)
        f.value = Path(s['fichier']).name
        f.font = F_SRC
        f.alignment = HAUT
        f.border = BAS
        o = sorted(s['valeurs'])
        gros = max(abs(v) for v in s['valeurs'].values())
        fmt = '#,##0' if gros >= 1000 else '#,##0.00'
        for i, p in enumerate(periodes):
            v = s['valeurs'].get(p)
            if v is not None:
                cel = ws.cell(5 + i, cc)
                cel.value = v
                cel.number_format = fmt
    return col0 + n


def main():
    if not SOURCE.exists():
        sys.exit('classeur _2 introuvable')
    print('Lecture de %s ...' % SOURCE.name)
    wb = openpyxl.load_workbook(SOURCE)
    vus = deja_utilisees(wb)
    print('  %d references de series deja mobilisees' % len(vus))

    print('Indexation du data lake ...')
    lake = charger_lake(ROOT) + lire_ipai()
    print('  %d series (dont IPAI en format long)' % len(lake))

    # filtrage + dedoublonnage (on conserve les libelles distincts)
    garde = {}
    for s in lake:
        if any(x in sansacc(s['fichier']) for x in EXCLURE_FICHIER):
            continue
        if any(re.match(p, sansacc(s['serie']).strip()) for p in EXCLURE_SERIE):
            continue
        o = sorted(s['valeurs'])
        # on accepte les series courtes si elles sont recentes : ce sont des
        # indicateurs recemment crees, qui s'allongeront
        if len(o) < 24 and not (len(o) >= 8 and int(str(o[-1])[:4]) >= 2025):
            continue
        if Path(s['fichier']).name.startswith('PIB trimestriel'):
            continue                     # traitees a part (cibles alternatives)
        cle = (s['freq'], sansacc(s['serie']),
               tuple(sorted((a, round(v, 6)) for a, v in s['valeurs'].items())))
        garde.setdefault(cle, s)
    candidats = [s for s in garde.values()
                 if (Path(s['fichier']).name, s['serie']) not in vus
                 and s['serie'] not in vus]
    print('  %d series candidates (non deja presentes dans _2)' % len(candidats))

    # classement
    par_branche = defaultdict(list)
    hors = []
    for s in candidats:
        if Path(s['fichier']).name == VA_FICHIER:
            continue                       # traite a part
        br = classer(s['serie'], s['fichier'])
        (par_branche[br] if br else hors).append(s)

    # VA retropolee, ajoutee comme cible etendue
    va = {sansacc(s['serie'])[:30]: s for s in lake
          if Path(s['fichier']).name == VA_FICHIER}

    # cibles alternatives : autres variantes du PIB trimestriel
    alt_pib = [s for s in lake
               if Path(s['fichier']).name.startswith('PIB trimestriel')
               and Path(s['fichier']).name != VA_FICHIER
               and len(s['valeurs']) >= 20]
    print('  %d series de PIB trimestriel disponibles comme cibles alternatives'
          % len(alt_pib))

    # bloc annuel
    annuel = lire_panel_annuel()
    print('  %d series annuelles / de campagne exploitables' % len(annuel))

    per_t_all, per_m_all = periodes_trim(), periodes_mois()
    ajouts = []

    for nom, lib_va in BRANCHES:
        if nom[:31] not in wb.sheetnames:
            print('  !! feuille absente : %s' % nom)
            continue
        ws = wb[nom[:31]]
        depart = ws.max_column + 2

        lst = sorted(par_branche.get(nom, []), key=lambda s: -len(s['valeurs']))
        trim = [s for s in lst if 'rimestr' in s['freq']]
        mens = [s for s in lst if 'ensuel' in s['freq']]

        cible = None
        for k, s in va.items():
            if k.startswith(sansacc(lib_va)[:24]):
                cible = dict(s)
                cible['serie'] = 'VA %s — base 2014 rétropolée (MDH)' % nom
                break
        if cible:
            trim.insert(0, cible)

        ws.cell(2, depart).value = ('AJOUTS issus du data lake — non presents dans '
                                    'la version precedente')
        ws.cell(2, depart).font = F_TITRE

        def etendue(series, ref):
            vusp = set()
            for s in series:
                vusp |= set(s['valeurs'])
            return [p for p in ref if p in vusp]

        fin = depart
        if trim:
            fin = ecrire_bloc(ws, depart, 'AJOUTS — INDICATEURS TRIMESTRIELS',
                              BL_T, H_T, 'Trimestre', etendue(trim, per_t_all), trim)
            for s in trim:
                o = sorted(s['valeurs'])
                ajouts.append([nom, s['serie'], 'trimestrielle',
                               s['fichier'], o[0], o[-1], len(o)])
        if mens:
            c0 = fin + 2
            fin = ecrire_bloc(ws, c0, 'AJOUTS — INDICATEURS MENSUELS',
                              BL_M, H_M, 'Mois', etendue(mens, per_m_all), mens)
            for s in mens:
                o = sorted(s['valeurs'])
                ajouts.append([nom, s['serie'], 'mensuelle',
                               s['fichier'], o[0], o[-1], len(o)])

        # ---- cibles alternatives (autres variantes du PIB trimestriel)
        cibles = []
        for s in alt_pib:
            if sansacc(s['serie']).startswith(sansacc(lib_va)[:18]) \
                    or classer(s['serie'], s['fichier']) == nom:
                d = dict(s)
                d['serie'] = '%s [%s]' % (s['serie'],
                                          variante_pib(Path(s['fichier']).name))
                cibles.append(d)
        if cibles:
            c0 = fin + 2
            fin = ecrire_bloc(ws, c0, 'CIBLES ALTERNATIVES (PIB trimestriel)',
                              BL_C, H_C, 'Trimestre',
                              etendue(cibles, per_t_all), cibles)
            for s in cibles:
                o = sorted(s['valeurs'])
                ajouts.append([nom, s['serie'], 'trimestrielle (cible alt.)',
                               s['fichier'], o[0], o[-1], len(o)])

        # ---- bloc annuel (calage, desagregation temporelle)
        ann = [s for s in annuel if classer(s['serie'], s['fichier']) == nom]
        ann.sort(key=lambda s: (-max(s['valeurs']), -len(s['valeurs'])))
        trop = max(0, len(ann) - MAX_ANNUEL)
        ann = ann[:MAX_ANNUEL]
        if ann:
            c0 = fin + 2
            annees = sorted({a for s in ann for a in s['valeurs']})
            fin = ecrire_bloc(ws, c0, 'BLOC ANNUEL — calage et désagrégation'
                              + (' (%d autres séries non retenues)' % trop if trop else ''),
                              BL_A, H_A, 'Année', annees, ann)
            for s in ann:
                o = sorted(s['valeurs'])
                ajouts.append([nom, s['serie'],
                               'annuelle' if s['freq'] == 'Annuelle' else 'campagne',
                               s['fichier'], o[0], o[-1], len(o)])
        ws.row_dimensions[3].height = 46
        ws.row_dimensions[4].height = 26
        print('  %-30s +%3d trimestriels, +%3d mensuels'
              % (nom, len(trim), len(mens)))

    # feuille de documentation des ajouts
    if 'Ajouts' in wb.sheetnames:
        del wb['Ajouts']
    wa = wb.create_sheet('Ajouts')
    wa.append(['Séries ajoutées depuis le data lake'])
    wa.cell(1, 1).font = Font(bold=True, size=14, color='7030A0')
    wa.append(['Le contenu d’origine du classeur est inchangé. '
               'Ces colonnes sont ajoutées à droite de chaque feuille.'])
    wa.cell(2, 1).font = Font(italic=True, size=10, color='595959')
    wa.append([])
    wa.append(['Branche', 'Série ajoutée', 'Fréquence', 'Fichier source',
               'Début', 'Fin', 'Nb obs'])
    for l in ajouts:
        wa.append(l)
    en = PatternFill('solid', fgColor='7030A0')
    for c, w in zip('ABCDEFG', (26, 50, 14, 58, 11, 11, 9)):
        wa.column_dimensions[c].width = w
        cel = wa.cell(4, ord(c) - 64)
        cel.font = Font(bold=True, color='FFFFFF')
        cel.fill = en
        cel.alignment = CENTRE
    wa.row_dimensions[4].height = 28
    wa.freeze_panes = 'A5'
    wa.auto_filter.ref = 'A4:G%d' % wa.max_row
    wa.sheet_view.showGridLines = False

    if hors:
        wh = wb.create_sheet('Ajouts non classés')
        wh.append(['Fichier source', 'Série', 'Fréquence', 'Début', 'Fin', 'Nb obs'])
        for s in hors:
            o = sorted(s['valeurs'])
            wh.append([s['fichier'], s['serie'], s['freq'], o[0], o[-1], len(o)])
        for c, w in zip('ABCDEF', (58, 46, 14, 11, 11, 9)):
            wh.column_dimensions[c].width = w

    wb.save(SORTIE)
    print('\n%s' % SORTIE.name)
    print('  %d colonnes ajoutees, %d non classees' % (len(ajouts), len(hors)))


if __name__ == '__main__':
    main()
