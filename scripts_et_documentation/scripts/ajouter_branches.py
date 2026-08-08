# -*- coding: utf-8 -*-
"""
Complete Etude_sectorielle_Maroc_2_complete.xlsx :

1. Ajoute les 4 branches absentes des comptes nationaux (28,2 % de la VA) et
   la feuille "PIB et impôts" (impots nets de subventions + PIB chaine), sans
   lesquels la somme des VA ne reconstitue pas le PIB.
2. Reconstruit integralement la feuille Metadonnees sur un format homogene :
   memes intitules de frequence, memes formats de periode, un role explicite
   par colonne, et la provenance complete.
3. Refait le Sommaire.

Le fichier est modifie sur place.
"""

import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parent))
from catalogue_lake import charger_lake                                  # noqa: E402
from completer_classeur import (lire_panel_annuel, variante_pib,         # noqa: E402
                                ecrire_bloc, BL_T, BL_M, BL_C, BL_A,
                                H_T, H_M, H_C, H_A, MAX_ANNUEL)
from enrichir_classeur import classer, sansacc, periodes_trim            # noqa: E402
from producteurs import Producteurs                                      # noqa: E402

CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'
VA_FICHIER = 'PIB trimestriel_PIB base2014 rétropolé 28 branc.csv'

# nouvelles feuilles : (nom de feuille, libelle de la VA dans les comptes)
NOUVELLES = [
    ('Administration publique',   'Administration publique et défense'),
    ('Éducation-santé',           'Education, santé humaine'),
    ('Services aux entreprises',  'Recherches et développement et services rendus'),
    ('Autres services',           'Autres services'),
]
POSTES_PIB = [('PIB et impôts', ['impots net subventions', 'PIB chainé'])]

EN = PatternFill('solid', fgColor='2F5597')
F_EN = Font(bold=True, color='FFFFFF', size=10)
F_TITRE = Font(bold=True, size=14, color='1F3864')
F_SOUS = Font(italic=True, size=10, color='595959')
CENTRE = Alignment(horizontal='center', vertical='center', wrap_text=True)


def norm(s):
    s = unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]', '', s.lower())


def styler(ws, ligne_en, largeurs, centrees=(), gel=None):
    """Mise en forme unique de toutes les feuilles de service :
    en-tete visible, volets figes, filtres, largeurs, pas de quadrillage."""
    for c, w in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cel = ws.cell(ligne_en, c)
        cel.font = F_EN
        cel.fill = EN
        cel.alignment = CENTRE
        cel.border = Border(bottom=Side(style='medium', color='1F3864'))
    ws.row_dimensions[ligne_en].height = 34
    for r in range(ligne_en + 1, ws.max_row + 1):
        for c in centrees:
            ws.cell(r, c).alignment = Alignment(horizontal='center')
    ws.freeze_panes = gel or ws.cell(ligne_en + 1, 1)
    ws.auto_filter.ref = 'A%d:%s%d' % (ligne_en,
                                       get_column_letter(len(largeurs)),
                                       ws.max_row)
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------- normalisation
def freq_std(v):
    f = sansacc(v)
    if 'cible alt' in f:
        return 'Trimestrielle'
    if 'campagne' in f or 'cheval' in f:
        return 'Campagne agricole'
    if 'trimestr' in f:
        return 'Trimestrielle'
    if 'mensuel' in f:
        return 'Mensuelle'
    if 'semestr' in f:
        return 'Semestrielle'
    if 'annuel' in f:
        return 'Annuelle'
    return '—'


def periode_std(v, freq):
    """Uniformise : 2016T1 / 2016M03 / 2016 / 1998-1999."""
    if v is None:
        return None
    s = str(v).strip()
    m = re.fullmatch(r'T([1-4])[-/](\d{4})', s)
    if m:
        return f'{m.group(2)}T{m.group(1)}'
    if re.fullmatch(r'\d{4}T[1-4]', s) or re.fullmatch(r'\d{4}M\d{2}', s):
        return s
    m = re.fullmatch(r'([a-zûéèA-Z]+)-(\d{2})', s)
    if m:
        mois = {'janv': 1, 'fevr': 2, 'févr': 2, 'mars': 3, 'avr': 4, 'mai': 5,
                'juin': 6, 'juil': 7, 'aout': 8, 'août': 8, 'sept': 9,
                'oct': 10, 'nov': 11, 'dec': 12, 'déc': 12}.get(m.group(1).lower()[:4])
        if mois:
            a = int(m.group(2))
            a += 2000 if a < 80 else 1900
            return f'{a}M{mois:02d}'
    if re.fullmatch(r'\d{4}(\.0)?', s):
        return s[:4]
    if re.fullmatch(r'\d{2,4}/\d{2,4}', s):
        return s
    return s


def freq_depuis_periode(p):
    """Deduit la frequence du FORMAT de la periode, seul indice fiable pour
    les colonnes d'origine (leur libelle ne porte pas la frequence)."""
    s = str(p or '').strip()
    if re.fullmatch(r'\d{4}T[1-4]', s):
        return 'Trimestrielle'
    if re.fullmatch(r'\d{4}M\d{2}', s):
        return 'Mensuelle'
    if re.fullmatch(r'\d{2,4}/\d{2,4}', s):
        return 'Campagne agricole'
    if re.fullmatch(r'\d{4}', s):
        return 'Annuelle'
    return '—'


def annee_de(p):
    m = re.search(r'(19|20)\d{2}', str(p) or '')
    return int(m.group(0)) if m else None


def series_bases_consolidees():
    """Libelles presents dans BDD SECTORIEL_MENSUEL / _TRIM.

    Ces bases sont le fichier DONT LES VALEURS ONT ETE RECOPIEES pour les
    colonnes restees sans source primaire : ce n'est pas une source au sens
    propre (elles ne documentent rien), mais c'est la provenance reelle de la
    recopie, et il vaut mieux l'ecrire que laisser un tiret."""
    import csv as _csv
    out = {}
    base = ROOT / 'Bases consolidees (multi-sources)'
    for nom in ('BDD SECTORIEL_MENSUEL.csv', 'BDD SECTORIEL_TRIM.csv'):
        p = base / nom
        if not p.exists():
            continue
        tete = open(p, encoding='utf-8-sig', errors='replace').readline()
        sep = ';' if tete.count(';') >= tete.count(',') else ','
        for r in _csv.reader(open(p, encoding='utf-8-sig', errors='replace'),
                             delimiter=sep):
            if r and r[0].strip():
                out.setdefault(norm(r[0])[:34],
                               str(Path('Bases consolidees (multi-sources)') / nom))
    return out


# ---------------------------------------------------------------- feuilles
def creer_feuilles(wb, lake, annuel, alt_pib, va):
    ajouts = []
    per_t = periodes_trim()

    def etendue(series, ref):
        vus = set()
        for s in series:
            vus |= set(s['valeurs'])
        return [p for p in ref if p in vus]

    for nom, lib in NOUVELLES + [(n, None) for n, _ in POSTES_PIB]:
        if nom in wb.sheetnames:
            del wb[nom]
        ws = wb.create_sheet(nom[:31])
        ws.cell(1, 1).value = 'BRANCHE — %s' % nom
        ws.cell(1, 1).font = F_TITRE
        ws.sheet_view.showGridLines = False

        if lib is None:                       # feuille PIB et impots
            libs = POSTES_PIB[0][1]
            cibles = [dict(s, serie=s['serie']) for s in lake
                      if Path(s['fichier']).name == VA_FICHIER
                      and s['serie'] in libs]
            ws.cell(2, 1).value = ('Impôts nets de subventions et PIB chaîné : '
                                   'PIB = Σ valeurs ajoutées + impôts nets.')
            ws.cell(2, 1).font = F_SOUS
        else:
            cibles = []
            for s in va.values():
                if sansacc(s['serie']).startswith(sansacc(lib)[:20]):
                    d = dict(s)
                    d['serie'] = 'VA %s — base 2014 rétropolée (MDH)' % nom
                    cibles.append(d)
                    break
            ws.cell(2, 1).value = ('Branche des comptes nationaux absente de la '
                                   'version précédente du classeur.')
            ws.cell(2, 1).font = F_SOUS

        fin = 1
        if cibles:
            fin = ecrire_bloc(ws, 1, 'VARIABLE CIBLE', BL_T, H_T, 'Trimestre',
                              etendue(cibles, per_t), cibles)
            for s in cibles:
                o = sorted(s['valeurs'])
                ajouts.append([nom, s['serie'], 'trimestrielle',
                               s['fichier'], o[0], o[-1], len(o)])

        # cibles alternatives
        alt = []
        for s in alt_pib:
            base = lib or ''
            if base and sansacc(s['serie']).startswith(sansacc(base)[:18]):
                d = dict(s)
                d['serie'] = '%s [%s]' % (s['serie'],
                                          variante_pib(Path(s['fichier']).name))
                alt.append(d)
        if alt:
            c0 = fin + 2
            fin = ecrire_bloc(ws, c0, 'CIBLES ALTERNATIVES (PIB trimestriel)',
                              BL_C, H_C, 'Trimestre', etendue(alt, per_t), alt)
            for s in alt:
                o = sorted(s['valeurs'])
                ajouts.append([nom, s['serie'], 'trimestrielle (cible alt.)',
                               s['fichier'], o[0], o[-1], len(o)])

        # indicateurs et annuel eventuels
        ind = [s for s in annuel if classer(s['serie'], s['fichier']) == nom]
        ind.sort(key=lambda s: (-max(s['valeurs']), -len(s['valeurs'])))
        ind = ind[:MAX_ANNUEL]
        if ind:
            c0 = fin + 2
            annees = sorted({a for s in ind for a in s['valeurs']})
            ecrire_bloc(ws, c0, 'BLOC ANNUEL — calage et désagrégation',
                        BL_A, H_A, 'Année', annees, ind)
            for s in ind:
                o = sorted(s['valeurs'])
                ajouts.append([nom, s['serie'],
                               'annuelle' if s['freq'] == 'Annuelle' else 'campagne',
                               s['fichier'], o[0], o[-1], len(o)])
        else:
            ws.cell(3, fin + 2).value = 'Aucun indicateur disponible dans le data lake'
            ws.cell(3, fin + 2).font = F_SOUS

        ws.row_dimensions[3].height = 46
        ws.row_dimensions[4].height = 26
        ws.column_dimensions['A'].width = 11
        ws.freeze_panes = 'B5'
        print('  + %-28s %d cible(s), %d alt., %d annuelles'
              % (nom, len(cibles), len(alt), len(ind)))
    return ajouts


# ---------------------------------------------------------------- metadonnees
def refaire_metadonnees(wb, lignes_ajouts):
    """Reconstruit la feuille a partir de Audit (colonnes d'origine) et
    Ajouts (series du lac), sur un format unique."""
    wa = wb['Audit']
    waj = wb['Ajouts']
    if 'Métadonnées' in wb.sheetnames:
        del wb['Métadonnées']
    wm = wb.create_sheet('Métadonnées', 1)

    prod = Producteurs()
    bases_cons = series_bases_consolidees()
    # sources telles que declarees dans le classeur d'origine (colonne "Source")
    declarees = {}
    orig = ROOT / 'Etude_sectorielle_Maroc_2.xlsx'
    if orig.exists():
        wo = openpyxl.load_workbook(orig, data_only=True)
        if 'Métadonnées' in wo.sheetnames:
            for r in wo['Métadonnées'].iter_rows(min_row=2, values_only=True):
                if r and r[0] and r[1]:
                    declarees[(norm(r[0]), norm(r[1])[:34])] = r[2]

    entetes = ['Branche', 'Colonne dans le classeur', 'Rôle', 'Fréquence',
               'Début', 'Fin', 'Année début', 'Année fin', 'Nb obs',
               'Source (institution productrice)', 'Attribution de la source',
               'Fichier source (chemin exact)', 'Série dans le fichier',
               'Origine', 'Alignement vérifié']
    # En-tetes en LIGNE 1 : c'est la seule position ou Excel les garde
    # toujours visibles (volets figes + filtres). Les explications sont
    # renvoyees dans la feuille "Lisez-moi".
    wm.append(entetes)

    n_org = n_aj = 0
    for r in wa.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        feuille, colonne, fichier, serie, statut = r[0], r[1], r[2], r[3], r[4]
        # Audit : [Feuille, Colonne, Fichier, Série, Statut,
        #          Nb valeurs écrites, Début source, Fin source, Remarque]
        nobs, deb, fin = r[5], r[6], r[7]
        role = ('Variable cible' if str(statut) == 'conservee'
                else 'Indicateur')
        d, fi = periode_std(deb, None), periode_std(fin, None)
        f = freq_depuis_periode(d)
        align = ('non vérifié' if str(statut).startswith('NON')
                 else 'oui — réaligné' if str(statut).startswith('REALIGNEE')
                 else 'oui')
        decl = declarees.get((norm(feuille), norm(colonne)[:34]))
        inst, attrib = prod.pour(fichier, serie or colonne, decl)
        fic = fichier if fichier not in (None, '?', '') else None
        ser_src = serie
        if fic is None:
            repris = bases_cons.get(norm(colonne)[:34])
            if repris:
                fic = '%s  ⚠ base consolidée, source primaire inconnue' % repris
                ser_src = colonne
            else:
                fic = '— origine inconnue'
                ser_src = '—'
        wm.append([feuille, colonne, role, f,
                   d, fi, annee_de(d), annee_de(fi), nobs,
                   inst, attrib, fic, ser_src, 'Colonne d’origine', align])
        n_org += 1

    for r in list(waj.iter_rows(min_row=5, values_only=True)) + lignes_ajouts:
        if not r or not r[0]:
            continue
        branche, serie, freq, fichier, deb, fin, nobs = r[:7]
        f = freq_std(freq)
        role = ('Cible alternative' if 'cible alt' in sansacc(freq)
                else 'Variable cible'
                if (str(serie).startswith('VA ') or branche == 'PIB et impôts')
                else 'Indicateur')
        d, fi = periode_std(deb, f), periode_std(fin, f)
        inst, attrib = prod.pour(fichier, serie)
        wm.append([branche, serie, role, f, d, fi, annee_de(d), annee_de(fi),
                   nobs, inst, attrib, fichier, serie, 'Ajout data lake', 'oui'])
        n_aj += 1

    # mise en forme
    styler(wm, 1, (24, 48, 17, 17, 11, 11, 11, 10, 8, 46, 30, 60, 42, 18, 17),
           centrees=(3, 4, 5, 6, 7, 8, 9, 14, 15), gel='C2')
    for r in range(2, wm.max_row + 1):
        if wm.cell(r, 15).value == 'non vérifié':
            wm.cell(r, 15).fill = PatternFill('solid', fgColor='FFC7CE')
        if wm.cell(r, 3).value in ('Variable cible', 'Cible alternative'):
            wm.cell(r, 3).fill = PatternFill('solid', fgColor='FCE4D6')
        if str(wm.cell(r, 10).value).startswith('—'):
            wm.cell(r, 10).fill = PatternFill('solid', fgColor='FFF2CC')
        if wm.cell(r, 11).value == 'communiqué par l’auteur de la base':
            wm.cell(r, 11).fill = PatternFill('solid', fgColor='DDEBF7')
    wm.sheet_view.showGridLines = False
    print('Metadonnees : %d colonnes d’origine + %d ajouts = %d lignes'
          % (n_org, n_aj, n_org + n_aj))
    return wm


def refaire_sommaire(wb, wm):
    if 'Sommaire' in wb.sheetnames:
        del wb['Sommaire']
    ws = wb.create_sheet('Sommaire', 0)
    ws.append(['Branche', 'Colonnes', 'dont cibles', 'dont indicateurs',
               'infra-annuels', 'alignement non vérifié'])

    stats = defaultdict(lambda: [0, 0, 0, 0, 0])
    for r in wm.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        s = stats[r[0]]
        s[0] += 1
        if r[2] in ('Variable cible', 'Cible alternative'):
            s[1] += 1
        else:
            s[2] += 1
        if r[3] in ('Trimestrielle', 'Mensuelle'):
            s[3] += 1
        if r[12] == 'non vérifié':
            s[4] += 1

    ordre = [f for f in wb.sheetnames
             if f not in ('Sommaire', 'Métadonnées', 'Audit', 'Ajouts',
                          'Ajouts non classés')]
    tot = [0, 0, 0, 0, 0]
    for b in ordre:
        s = stats.get(b, [0, 0, 0, 0, 0])
        ws.append([b] + s)
        rr = ws.max_row
        ws.cell(rr, 1).hyperlink = "#'%s'!A1" % b
        ws.cell(rr, 1).font = Font(color='0563C1', underline='single')
        for c in range(2, 7):
            ws.cell(rr, c).alignment = Alignment(horizontal='center')
        if s[4]:
            ws.cell(rr, 6).fill = PatternFill('solid', fgColor='FFC7CE')
        tot = [a + b_ for a, b_ in zip(tot, s)]
    ws.append(['TOTAL'] + tot)
    for c in range(1, 7):
        ws.cell(ws.max_row, c).font = Font(bold=True)
    styler(ws, 1, (32, 12, 13, 16, 15, 22), centrees=(2, 3, 4, 5, 6))
    print('Sommaire : %d feuilles, %d colonnes' % (len(ordre), tot[0]))


def uniformiser_services(wb):
    """Meme presentation pour Audit, Ajouts et Ajouts non classés :
    en-tete en ligne 1, visible, filtre et fige."""
    # Ajouts : l'en-tete etait en ligne 4, avec deux lignes de titre au-dessus
    if 'Ajouts' in wb.sheetnames:
        wj = wb['Ajouts']
        if str(wj.cell(1, 1).value or '').startswith('Séries ajoutées'):
            wj.delete_rows(1, 3)
        styler(wj, 1, (26, 50, 20, 58, 12, 12, 10), centrees=(3, 5, 6, 7))
    if 'Audit' in wb.sheetnames:
        styler(wb['Audit'], 1, (24, 50, 50, 40, 24, 12, 12, 12, 34),
               centrees=(6, 7, 8))
    if 'Ajouts non classés' in wb.sheetnames:
        styler(wb['Ajouts non classés'], 1, (58, 46, 16, 12, 12, 10),
               centrees=(3, 4, 5, 6))


def lisez_moi(wb):
    if 'Lisez-moi' in wb.sheetnames:
        del wb['Lisez-moi']
    ws = wb.create_sheet('Lisez-moi', 0)
    lignes = [
        ('ÉTUDE SECTORIELLE MAROC — tableau de synthèse', 'titre'),
        ('', ''),
        ('Objet', 'section'),
        ('Nowcasting du PIB marocain par branche d’activité (approche offre). '
         'Une feuille par branche des comptes nationaux, plus les impôts nets '
         'de subventions qui permettent de reconstituer le PIB.', ''),
        ('', ''),
        ('Comment lire une feuille de branche', 'section'),
        ('Ligne 3 : nom de la colonne · Ligne 4 : fichier source · '
         'Données à partir de la ligne 5.', ''),
        ('Les blocs sont identifiés par un bandeau de couleur : bleu = colonnes '
         'd’origine, violet = ajouts trimestriels, orange = ajouts mensuels, '
         'jaune = cibles alternatives, vert = bloc annuel.', ''),
        ('Chaque bloc a son propre axe de dates ; les séries sont jointes sur '
         'leur date, jamais collées par position.', ''),
        ('', ''),
        ('Feuilles de service', 'section'),
        ('Sommaire   : nombre de colonnes par branche, liens cliquables.', ''),
        ('Métadonnées: une ligne par colonne — rôle, fréquence, couverture, '
         'fichier source, série d’origine, statut d’alignement.', ''),
        ('Audit      : reconstruction des colonnes d’origine (colonnes '
         'réalignées, vérifiées, ou restées sans source).', ''),
        ('Ajouts     : séries ajoutées depuis le data lake.', ''),
        ('', ''),
        ('Réserve importante', 'section'),
        ('21 colonnes proviennent de la base consolidée BDD SECTORIEL, qui ne '
         'documente aucune source. Leur alignement n’a pas pu être vérifié, '
         'malgré deux tentatives contre l’index complet du data lake. '
         'Elles sont signalées « non vérifié » en rouge dans Métadonnées.', 'alerte'),
    ]
    for i, (txt, style) in enumerate(lignes, start=1):
        c = ws.cell(i, 1)
        c.value = txt
        if style == 'titre':
            c.font = Font(bold=True, size=16, color='1F3864')
        elif style == 'section':
            c.font = Font(bold=True, size=12, color='2F5597')
        elif style == 'alerte':
            c.font = Font(size=10, color='9C0006')
            c.fill = PatternFill('solid', fgColor='FFC7CE')
        else:
            c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 118
    for i in (4, 8, 9, 13, 14, 18):
        ws.row_dimensions[i].height = 32
    ws.sheet_view.showGridLines = False


def main():
    print('Lecture du classeur ...')
    wb = openpyxl.load_workbook(CLASSEUR)
    lake = charger_lake(ROOT)
    annuel = lire_panel_annuel()
    va = {sansacc(s['serie'])[:30]: s for s in lake
          if Path(s['fichier']).name == VA_FICHIER}
    alt_pib = [s for s in lake
               if Path(s['fichier']).name.startswith('PIB trimestriel')
               and Path(s['fichier']).name != VA_FICHIER
               and len(s['valeurs']) >= 20]

    print('Creation des feuilles manquantes ...')
    ajouts = creer_feuilles(wb, lake, annuel, alt_pib, va)

    wm = refaire_metadonnees(wb, ajouts)
    refaire_sommaire(wb, wm)
    uniformiser_services(wb)
    lisez_moi(wb)
    wb.save(CLASSEUR)
    print('\n%s mis a jour' % CLASSEUR.name)


if __name__ == '__main__':
    main()
