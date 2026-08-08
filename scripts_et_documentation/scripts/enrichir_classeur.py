# -*- coding: utf-8 -*-
"""
Construit Etude_sectorielle_Maroc_3.xlsx : toutes les series exploitables du
data lake, classees par branche, datees et sourcees.

Principes
---------
* Variable cible : VA trimestrielle rétropolée base 2014, 28 branches
  (1998T1 -> 2026T1, 113 trimestres) au lieu des 49 trimestres actuels.
* Chaque serie est jointe SUR SA DATE, jamais collee par position.
* Chaque colonne porte son fichier source et le nom exact de la serie dans ce
  fichier : la provenance est verifiable ligne par ligne.
* Les bases consolidees intermediaires (BDD SECTORIEL_*) sont volontairement
  ecartees : elles ne documentent aucune source et leur alignement est
  invérifiable.

    python enrichir_classeur.py
"""

import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parent))
from catalogue_lake import charger_lake                      # noqa: E402

SORTIE = ROOT / 'Etude_sectorielle_Maroc_3.xlsx'
FICHIER_VA = 'PIB trimestriel_PIB base2014 rétropolé 28 branc.csv'

AN_MIN_T, AN_MAX_T = 1998, 2026
AN_MIN_M, AN_MAX_M = 1998, 2026

# --- branches et variable cible -------------------------------------------
BRANCHES = [
    ('Agriculture',                 'Agriculture'),
    ('Pêche',                       'Pêche'),
    ("Industrie d'extraction",      'Industrie d’extraction'),
    ('Industrie de transformation', 'Industrie de transformation'),
    ('Électricité, gaz, eau',       'Distribution d’électricité et de gaz'),
    ('Construction',                'Construction'),
    ('Commerce',                    'Commerce de gros et de détail'),
    ('Transports',                  'Transports et entreposage'),
    ('Hébergement-restauration',    'Activités d’hébergement et de restauration'),
    ('Information-communication',   'Information et communication'),
    ('Finances et assurances',      'Activités financières et d’assurances'),
    ('Immobilier',                  'Activités immobilières'),
]

# --- regles de classement --------------------------------------------------
MOTS = {
    'Pêche': ['peche', 'debarquement', 'halieutique', 'flotte', 'poisson',
              'cephalopod', 'crustac', 'coquillage', 'algues', 'thon',
              'sardine', 'oursin'],
    'Agriculture': ['precipitation', 'pluviom', 'barrage', 'cereale', 'culture',
                    'rendement', 'cheptel', 'abattage', 'superficie', 'agricole',
                    'climat', 'oleagineux', 'betterave', 'canne', 'agrume',
                    'olive', 'elevage', 'miel', 'oeufs', 'viande', 'lait'],
    "Industrie d'extraction": ['phosphate', 'miniere', 'minier', 'extractive',
                               'extraction', 'ocp', 'soufre', 'barytine',
                               'plomb', 'zinc', 'fluorine', 'manganese'],
    'Industrie de transformation': ['industriel', 'industrie', 'ipi', 'opinion',
                                    'conjoncture', 'capacite', 'manufactur',
                                    'sucre', 'sucriere', 'minoterie', 'laitiere',
                                    'oleicole', 'artisan', 'ecrasement',
                                    'entreprises industrielles'],
    'Électricité, gaz, eau': ['energie', 'electricit', 'onee', 'one-be', 'gaz',
                              'combustible', 'eau potable', 'petrolier',
                              'raffin', 'brent', 'butane', 'fioul', 'charbon',
                              'eolien', 'solaire', 'hydraulique', 'thermique',
                              'gwh', 'step'],
    'Construction': ['ciment', 'construction', 'logement', 'autorisation',
                     'batie', 'plancher', 'piece'],
    'Commerce': ['commerce exterieur', 'exportation', 'importation',
                 'commerce', 'ide', 'produits finis', 'demi produits'],
    'Transports': ['transport', 'trafic', 'portuaire', 'aerien', 'ferroviaire',
                   'vehicule', 'routier', 'immatriculation', 'wagon',
                   'voyageur', 'tonnage', 'accident'],
    'Hébergement-restauration': ['touris', 'hotel', 'nuitee', 'arrivee',
                                 'arrives', 'occupation', 'hebergement',
                                 'mre', 'lits'],
    'Information-communication': ['telecom', 'anrt', 'internet', 'mobile',
                                  'telephonie', 'sms', 'arpm', 'domaine',
                                  'poste', 'e-commerce', 'bande passante'],
    'Finances et assurances': ['credit', 'monetaire', 'depot', 'reserve',
                               'assurance', 'prime', 'bancaire', 'm3',
                               'souffrance', 'echeance', 'tresorerie'],
    'Immobilier': ['ipai', 'immobilier', 'habitat', 'foncier', 'appartement',
                   'villa', 'terrain'],
}

# fichiers a ne pas mobiliser
EXCLURE_FICHIER = ['bdd sectoriel', 'etude_sectorielle', 'manar_panel',
                   '_transfert', 'graphify-out']
EXCLURE_SERIE = [r'^v\s*(\(\d+\))?$', r'^vl\s*(\(\d+\))?$', r'^unnamed',
                 r'^colonne\d*$', r'^[a-z]$']


def sansacc(s):
    s = unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode()
    return s.lower()


def classer(serie, fichier):
    """-> nom de branche ou None.

    On n'utilise que le NOM du fichier, jamais son chemin : le dossier
    "racine (comptes nationaux, sectoriel, conjoncture, prix)" contient le mot
    "conjoncture", qui rattachait a tort tous ses fichiers a l'industrie.
    """
    t_ser, t_fic = sansacc(serie), sansacc(Path(fichier).name)
    meilleur, score_max = None, 0
    for source, poids in ((t_ser, 3), (t_fic, 1)):
        for br, mots in MOTS.items():
            sc = sum(poids * len(m) for m in mots if m in source)
            if sc > score_max:
                meilleur, score_max = br, sc
    return meilleur


def periodes_trim():
    return [f'{a}T{t}' for a in range(AN_MIN_T, AN_MAX_T + 1) for t in (1, 2, 3, 4)]


def periodes_mois():
    return [f'{a}M{m:02d}' for a in range(AN_MIN_M, AN_MAX_M + 1)
            for m in range(1, 13)]


def main():
    print('Indexation du data lake ...')
    lake = charger_lake(ROOT)
    print('  %d series brutes' % len(lake))

    # --- filtrage -----------------------------------------------------------
    retenues, signatures = [], {}
    for s in lake:
        fic, ser = s['fichier'], s['serie']
        if any(x in sansacc(fic) for x in EXCLURE_FICHIER):
            continue
        if any(re.match(p, sansacc(ser).strip()) for p in EXCLURE_SERIE):
            continue
        if len(s['valeurs']) < 24:
            continue
        # doublons : meme contenu -> on garde le fichier au nom le plus simple
        sig = (s['freq'], tuple(sorted((k, round(v, 6))
                                       for k, v in s['valeurs'].items())))
        garde = signatures.get(sig)
        if garde is None or len(fic) < len(garde['fichier']):
            signatures[sig] = s
    retenues = list(signatures.values())
    print('  %d series apres filtrage et dedoublonnage' % len(retenues))

    # --- variable cible -----------------------------------------------------
    va = {}
    for s in lake:
        if Path(s['fichier']).name == FICHIER_VA:
            va[sansacc(s['serie'])[:30]] = s
    print('  %d series de VA retropolee disponibles' % len(va))

    # --- classement ---------------------------------------------------------
    par_branche = defaultdict(list)
    non_classees = []
    for s in retenues:
        if Path(s['fichier']).name.startswith('PIB trimestriel'):
            continue                                   # reserve a la cible
        br = classer(s['serie'], s['fichier'])
        (par_branche[br] if br else non_classees).append(s)

    for nom, _ in BRANCHES:
        print('    %-30s %3d indicateurs' % (nom, len(par_branche.get(nom, []))))
    print('    %-30s %3d' % ('(non classees)', len(non_classees)))

    # --- construction du classeur ------------------------------------------
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sources = []

    # styles
    F_TITRE = Font(bold=True, size=15, color='1F3864')
    F_SOUS = Font(italic=True, size=10, color='595959')
    F_BLOC = Font(bold=True, size=11, color='FFFFFF')
    F_COL = Font(bold=True, size=10)
    F_SRC = Font(italic=True, size=8, color='7F7F7F')
    F_AXE = Font(bold=True, size=10, color='1F3864')
    BL_T = PatternFill('solid', fgColor='2F5597')      # bandeau trimestriel
    BL_M = PatternFill('solid', fgColor='548235')      # bandeau mensuel
    H_CIBLE = PatternFill('solid', fgColor='FBE2D5')
    H_T = PatternFill('solid', fgColor='DDEBF7')
    H_M = PatternFill('solid', fgColor='E2EFDA')
    AXE_F = PatternFill('solid', fgColor='F2F2F2')
    CENTRE = Alignment(horizontal='center', vertical='center', wrap_text=True)
    HAUT = Alignment(horizontal='center', vertical='top', wrap_text=True)
    bas = Border(bottom=Side(style='thin', color='BFBFBF'))

    L_BLOC, L_COL, L_FIC, L_SER, L_PER, L_DATA = 4, 5, 6, 7, 8, 9

    def ecrire_bloc(ws, col0, titre_bloc, fill_bandeau, axe_nom, periodes,
                    colonnes, fill_col, cible_idx=None):
        """colonnes : [(titre, serie_dict, est_cible)] ; renvoie derniere colonne."""
        n = len(colonnes)
        if n:
            ws.merge_cells(start_row=L_BLOC, start_column=col0,
                           end_row=L_BLOC, end_column=col0 + n)
        cel = ws.cell(L_BLOC, col0)
        cel.value = titre_bloc
        cel.font = F_BLOC
        cel.fill = fill_bandeau
        cel.alignment = CENTRE

        ws.cell(L_COL, col0).value = axe_nom
        ws.cell(L_COL, col0).font = F_AXE
        ws.cell(L_COL, col0).alignment = CENTRE
        ws.cell(L_COL, col0).fill = AXE_F
        for r in (L_FIC, L_SER, L_PER):
            ws.cell(r, col0).fill = AXE_F
        ws.cell(L_PER, col0).border = bas
        for i, p in enumerate(periodes):
            c = ws.cell(L_DATA + i, col0)
            c.value = p
            c.font = F_AXE
            c.alignment = Alignment(horizontal='center')
            c.fill = AXE_F

        for j, (titre, s, est_cible) in enumerate(colonnes, start=1):
            c = col0 + j
            lettre = get_column_letter(c)
            ws.column_dimensions[lettre].width = 15

            t = ws.cell(L_COL, c)
            t.value = titre
            t.font = F_COL
            t.fill = H_CIBLE if est_cible else fill_col
            t.alignment = HAUT

            f = ws.cell(L_FIC, c)
            f.value = Path(s['fichier']).name
            f.font = F_SRC
            f.alignment = HAUT

            se = ws.cell(L_SER, c)
            se.value = s['serie']
            se.font = F_SRC
            se.alignment = HAUT

            o = sorted(s['valeurs'])
            p = ws.cell(L_PER, c)
            p.value = '%s → %s  (%d)' % (o[0], o[-1], len(o))
            p.font = F_SRC
            p.alignment = Alignment(horizontal='center')
            p.border = bas

            gros = max(abs(v) for v in s['valeurs'].values())
            fmt = '#,##0' if gros >= 1000 else '#,##0.00'
            for i, per in enumerate(periodes):
                v = s['valeurs'].get(per)
                if v is not None:
                    cel = ws.cell(L_DATA + i, c)
                    cel.value = v
                    cel.number_format = fmt
        return col0 + n

    per_t_all, per_m_all = periodes_trim(), periodes_mois()

    for nom, lib_va in BRANCHES:
        ws = wb.create_sheet(nom[:31])
        ind = par_branche.get(nom, [])
        trim = sorted([s for s in ind if 'rimestr' in s['freq']],
                      key=lambda s: -len(s['valeurs']))
        mens = sorted([s for s in ind if 'ensuel' in s['freq']],
                      key=lambda s: -len(s['valeurs']))

        cible = None
        for k, s in va.items():
            if k.startswith(sansacc(lib_va)[:24]):
                cible = s
                break

        colonnes_t = ([('VA %s — variable cible (MDH)' % nom, cible, True)]
                      if cible else [])
        colonnes_t += [(s['serie'], s, False) for s in trim]
        colonnes_m = [(s['serie'], s, False) for s in mens]

        # axes reduits a l'etendue reellement couverte
        def etendue(cols, ref):
            vus = set()
            for _, s, _ in cols:
                vus |= set(s['valeurs'])
            dispo = [p for p in ref if p in vus]
            return dispo or ref[:1]

        per_t = etendue(colonnes_t, per_t_all) if colonnes_t else []
        per_m = etendue(colonnes_m, per_m_all) if colonnes_m else []

        # titre
        ws.cell(1, 1).value = 'BRANCHE — %s' % nom
        ws.cell(1, 1).font = F_TITRE
        ws.cell(2, 1).value = (
            '%d indicateur(s) trimestriel(s) · %d mensuel(s) · '
            'variable cible : %s' %
            (len(trim), len(mens),
             'VA base 2014 rétropolée, %s → %s' % (per_t[0], per_t[-1])
             if cible else 'ABSENTE'))
        ws.cell(2, 1).font = F_SOUS

        fin_t = ecrire_bloc(ws, 1, 'BLOC TRIMESTRIEL', BL_T, 'Trimestre',
                            per_t, colonnes_t, H_T) if colonnes_t else 1
        for titre, s, est_cible in colonnes_t:
            o = sorted(s['valeurs'])
            sources.append([nom, titre, 'trimestrielle', s['fichier'], s['serie'],
                            o[0], o[-1], len(o),
                            'variable cible' if est_cible else 'indicateur'])

        if colonnes_m:
            col0_m = fin_t + 2
            ecrire_bloc(ws, col0_m, 'BLOC MENSUEL', BL_M, 'Mois',
                        per_m, colonnes_m, H_M)
            ws.column_dimensions[get_column_letter(col0_m)].width = 11
            for titre, s, _ in colonnes_m:
                o = sorted(s['valeurs'])
                sources.append([nom, titre, 'mensuelle', s['fichier'], s['serie'],
                                o[0], o[-1], len(o), 'indicateur'])
        else:
            ws.cell(L_BLOC, fin_t + 2).value = 'Aucun indicateur mensuel disponible'
            ws.cell(L_BLOC, fin_t + 2).font = F_SOUS

        ws.column_dimensions['A'].width = 11
        ws.row_dimensions[L_COL].height = 46
        ws.row_dimensions[L_FIC].height = 30
        ws.row_dimensions[L_SER].height = 30
        ws.freeze_panes = ws.cell(L_DATA, 2)
        ws.sheet_view.showGridLines = False

    EN_TETE = PatternFill('solid', fgColor='2F5597')
    F_EN = Font(bold=True, color='FFFFFF', size=10)

    def styler_tableau(ws, ligne_entete, largeurs):
        for c, w in enumerate(largeurs, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
            cel = ws.cell(ligne_entete, c)
            cel.font = F_EN
            cel.fill = EN_TETE
            cel.alignment = Alignment(horizontal='center', vertical='center',
                                      wrap_text=True)
        ws.row_dimensions[ligne_entete].height = 30
        ws.freeze_panes = ws.cell(ligne_entete + 1, 1)
        ws.auto_filter.ref = 'A%d:%s%d' % (ligne_entete,
                                           get_column_letter(len(largeurs)),
                                           ws.max_row)
        ws.sheet_view.showGridLines = False

    # ---- feuille Sources ---------------------------------------------------
    wsrc = wb.create_sheet('Sources', 0)
    wsrc.append(['Provenance de chaque colonne du classeur'])
    wsrc.cell(1, 1).font = F_TITRE
    wsrc.append(['Chaque ligne permet de remonter une valeur jusqu’à son fichier '
                 'd’origine. Filtres actifs.'])
    wsrc.cell(2, 1).font = F_SOUS
    wsrc.append([])
    wsrc.append(['Branche', 'Colonne dans le classeur', 'Fréquence',
                 'Fichier source', 'Série dans le fichier', 'Début', 'Fin',
                 'Nb obs', 'Rôle'])
    for l in sources:
        wsrc.append(l)
    styler_tableau(wsrc, 4, (26, 44, 13, 58, 44, 10, 10, 9, 14))

    # ---- feuille Sommaire --------------------------------------------------
    wsom = wb.create_sheet('Sommaire', 0)
    wsom.append(['ÉTUDE SECTORIELLE MAROC — base de travail nowcasting'])
    wsom.cell(1, 1).font = Font(bold=True, size=16, color='1F3864')
    wsom.append(['Variable cible : valeur ajoutée trimestrielle, base 2014 '
                 'rétropolée 28 branches (1998T1 → 2026T1, 113 trimestres)'])
    wsom.cell(2, 1).font = F_SOUS
    wsom.append(['Toutes les séries sont jointes sur leur date et tracées '
                 'jusqu’à leur fichier d’origine (voir feuille Sources).'])
    wsom.cell(3, 1).font = F_SOUS
    wsom.append([])
    wsom.append(['Branche', 'Variable cible', 'Indicateurs trimestriels',
                 'Indicateurs mensuels', 'Total colonnes'])
    for nom, _ in BRANCHES:
        lignes = [s for s in sources if s[0] == nom]
        nt = sum(1 for s in lignes if s[2] == 'trimestrielle' and s[8] == 'indicateur')
        nm = sum(1 for s in lignes if s[2] == 'mensuelle')
        nc = sum(1 for s in lignes if s[8] == 'variable cible')
        wsom.append([nom, 'oui' if nc else 'ABSENTE', nt, nm, nt + nm + nc])
        r = wsom.max_row
        wsom.cell(r, 1).hyperlink = "#'%s'!A1" % nom[:31]
        wsom.cell(r, 1).font = Font(color='0563C1', underline='single')
        for c in (3, 4, 5):
            wsom.cell(r, c).alignment = Alignment(horizontal='center')
        if nt + nm == 0:
            for c in range(1, 6):
                wsom.cell(r, c).fill = PatternFill('solid', fgColor='FFC7CE')
    wsom.append(['TOTAL', '', sum(1 for s in sources if s[2] == 'trimestrielle'
                                  and s[8] == 'indicateur'),
                 sum(1 for s in sources if s[2] == 'mensuelle'), len(sources)])
    for c in range(1, 6):
        wsom.cell(wsom.max_row, c).font = Font(bold=True)
    styler_tableau(wsom, 5, (32, 16, 24, 22, 16))

    # ---- feuille des non classees -----------------------------------------
    wnc = wb.create_sheet('Non classées')
    wnc.append(['Séries non rattachées automatiquement à une branche'])
    wnc.cell(1, 1).font = F_TITRE
    wnc.append(['À rattacher manuellement si elles présentent un intérêt.'])
    wnc.cell(2, 1).font = F_SOUS
    wnc.append([])
    wnc.append(['Fichier source', 'Série', 'Fréquence', 'Début', 'Fin', 'Nb obs'])
    for s in sorted(non_classees, key=lambda x: x['fichier']):
        o = sorted(s['valeurs'])
        wnc.append([s['fichier'], s['serie'], s['freq'], o[0], o[-1], len(o)])
    styler_tableau(wnc, 4, (58, 44, 14, 11, 11, 9))

    wb.save(SORTIE)
    print('\n%s' % SORTIE.name)
    print('  %d colonnes sourcees, %d non classees' % (len(sources), len(non_classees)))


if __name__ == '__main__':
    main()
