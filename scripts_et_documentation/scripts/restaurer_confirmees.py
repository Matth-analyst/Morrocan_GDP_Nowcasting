# -*- coding: utf-8 -*-
"""
Reintegre, dans Etude_sectorielle_Maroc_2_complete.xlsx, les 10 colonnes
(sur les 31 retirees par corriger_fidelite.py) dont l'auteur de
BDD SECTORIEL_MENSUEL/TRIM a confirme la source institutionnelle :

  SIG Maroc         -> precipitations, temperature (Agriculture)
  ACAPS             -> primes, prestations d'assurance (Finances)
  Banque centrale   -> taux credits immobiliers (Immobilier),
                       depots a vue, masse monetaire M3,
                       avoirs officiels de reserve (Finances)
  HCP               -> indices import/export (Commerce)

Les 21 autres colonnes retirees (phosphates OCP, ANRT, IPAI niveau,
electricite, commerce Office des Changes...) restent retirees : aucune
source n'a ete confirmee pour elles.

Statut applique aux colonnes restaurees : "confirmé par l'auteur — date NON
vérifiée" (ni plus, ni moins que ce que l'on sait reellement).

Les valeurs sont reprises A L'IDENTIQUE de la version du classeur juste
avant leur retrait (commit git HEAD), pas re-derivees, pour eviter tout
risque de re-parsing.
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'
AVANT = Path(r'C:\Users\HP\AppData\Local\Temp\claude_restore\avant_retrait.xlsx')

# (feuille, titre exact) -> institution confirmee
A_RESTAURER = {
    ('Agriculture', 'Moyenne des precipitations'): 'SIG Maroc',
    ('Agriculture', 'Température_moyenne'): 'SIG Maroc',
    ('Finances et assurances', 'Primes versées (MDH)'):
        'ACAPS (Autorité de Contrôle des Assurances et de la Prévoyance Sociale)',
    ('Finances et assurances', 'Prestations (MDH)'):
        'ACAPS (Autorité de Contrôle des Assurances et de la Prévoyance Sociale)',
    ('Immobilier', 'taux Crédits immobiliers'): 'Bank Al-Maghrib',
    ('Finances et assurances', 'Dépôts à vue auprés des banques'): 'Bank Al-Maghrib',
    ('Finances et assurances', 'Masse monétaire (M3)'): 'Bank Al-Maghrib',
    ('Finances et assurances', 'Avoirs officiels de réserve'): 'Bank Al-Maghrib',
    ('Commerce', 'Indice IMPORTATIONS'): 'HCP (Haut-Commissariat au Plan)',
    ('Commerce', 'Indice EXPORTATIONS'): 'HCP (Haut-Commissariat au Plan)',
}

STATUT = 'confirmé par l’auteur — date non vérifiée'
ORANGE = PatternFill('solid', fgColor='FFE699')


def trouver_colonne(ws, titre):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(3, c).value or '').strip() == titre.strip():
            return c
    return None


def trouver_axe(ws, c):
    for cc in range(c, 0, -1):
        v3, v4 = ws.cell(3, cc).value, ws.cell(4, cc).value
        if v3 in ('Trimestre', 'Mois', 'Année') or v4 in ('Trimestre', 'Mois', 'Année'):
            return cc, (v3 if v3 in ('Trimestre', 'Mois', 'Année') else v4)
    return None, None


def cle(ws, r, col_axe, type_axe):
    """Cle de periode brute (peu importe le format, on recopie tel quel)."""
    return ws.cell(r, col_axe).value


def main():
    wb_avant = openpyxl.load_workbook(AVANT, data_only=True)
    wb = openpyxl.load_workbook(CLASSEUR)

    restaurees = []
    for (feuille, titre), inst in A_RESTAURER.items():
        ws_av = wb_avant[feuille]
        c_av = trouver_colonne(ws_av, titre)
        if c_av is None:
            print('  !! introuvable dans la version AVANT :', feuille, '|', titre)
            continue
        axe_av, type_axe = trouver_axe(ws_av, c_av)

        # relever (periode_brute, valeur) depuis la version d'avant
        paires = []
        for r in range(5, ws_av.max_row + 1):
            k = ws_av.cell(r, axe_av).value
            v = ws_av.cell(r, c_av).value
            if k is not None and isinstance(v, (int, float)):
                paires.append((k, v))
        if not paires:
            print('  !! aucune valeur trouvee pour :', feuille, '|', titre)
            continue

        # inserer une nouvelle colonne dans le classeur ACTUEL, a la suite
        # du bloc d'origine (juste apres la derniere colonne du bloc contenant
        # l'axe de meme type dans la meme feuille)
        ws = wb[feuille]
        # on cherche l'axe de meme type le plus proche dans le classeur actuel
        axe_actuel = None
        for c in range(1, ws.max_column + 1):
            v3, v4 = ws.cell(3, c).value, ws.cell(4, c).value
            if (v3 == type_axe or v4 == type_axe):
                axe_actuel = c
        if axe_actuel is None:
            print('  !! aucun axe %s trouve dans %s pour reinsertion' % (type_axe, feuille))
            continue
        # colonne d'insertion : juste apres la derniere colonne non vide du bloc
        c_ins = axe_actuel + 1
        while not all(ws.cell(r, c_ins).value in (None, '') for r in range(1, 12)):
            c_ins += 1
        ws.insert_cols(c_ins, 1)

        ws.cell(3, c_ins).value = titre
        ws.cell(4, c_ins).value = 'BDD SECTORIEL (source institutionnelle confirmée : %s)' % inst
        ws.cell(2, c_ins).value = 'RESTAUREE — %s' % STATUT
        ws.cell(3, c_ins).fill = ORANGE

        # ecrire les valeurs en reutilisant les lignes de l'axe actuel quand la
        # periode existe deja, sinon en etendant
        lignes_axe = {}
        for r in range(5, ws.max_row + 1):
            k = ws.cell(r, axe_actuel).value
            if k is not None:
                lignes_axe[k] = r
        derniere = max(lignes_axe.values()) if lignes_axe else 4
        for k, v in paires:
            if k not in lignes_axe:
                derniere += 1
                ws.cell(derniere, axe_actuel).value = k
                lignes_axe[k] = derniere
            ws.cell(lignes_axe[k], c_ins).value = v

        restaurees.append([feuille, titre, inst, len(paires)])
        print('  %-28s | %-46s | %-45s | %d valeurs'
              % (feuille, titre[:46], inst[:45], len(paires)))

    wb.save(CLASSEUR)
    print('\n%d colonnes restaurees, enregistrees dans %s' % (len(restaurees), CLASSEUR.name))
    return restaurees


if __name__ == '__main__':
    main()
