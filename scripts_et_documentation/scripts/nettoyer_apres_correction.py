# -*- coding: utf-8 -*-
"""
Met a jour Metadonnees, Audit et Sommaire apres corriger_fidelite.py :
- retire les lignes des 31 colonnes supprimees
- corrige les 2 lignes des colonnes remplacees (nouvelle source, statut)
- recalcule les compteurs du Sommaire
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parent))
from corriger_fidelite import A_RETIRER, A_REMPLACER   # noqa: E402

CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'
RETIRES_TITRES = {(f, t.strip()) for f, t in A_RETIRER}
REMPLACES_TITRES = {(f, t.strip()): (fic, ser, meth) for f, t, fic, ser, meth in A_REMPLACER}

F_EN = Font(bold=True, color='FFFFFF', size=10)
EN = PatternFill('solid', fgColor='2F5597')
VERT = PatternFill('solid', fgColor='C6EFCE')


def main():
    wb = openpyxl.load_workbook(CLASSEUR)

    # ---- Metadonnees --------------------------------------------------------
    wm = wb['Métadonnées']
    hdr = [c.value for c in wm[1]]
    idx_branche, idx_colonne = 0, 1
    idx_fic, idx_ser, idx_origine, idx_align = 11, 12, 13, 14

    a_supprimer = []
    n_corrige = 0
    for r in range(2, wm.max_row + 1):
        branche = wm.cell(r, 1).value
        colonne = str(wm.cell(r, 2).value or '').strip()
        if not branche:
            continue
        cle = (branche, colonne)
        if cle in RETIRES_TITRES:
            a_supprimer.append(r)
        elif cle in REMPLACES_TITRES:
            fic, ser, meth = REMPLACES_TITRES[cle]
            wm.cell(r, idx_fic + 1).value = fic
            wm.cell(r, idx_ser + 1).value = ser
            wm.cell(r, idx_origine + 1).value = 'Remplacée (source primaire confirmée)'
            wm.cell(r, idx_align + 1).value = 'oui'
            wm.cell(r, idx_align + 1).fill = VERT
            wm.cell(r, 10).value = 'Manar-Stat'   # institution productrice
            n_corrige += 1
    for r in sorted(a_supprimer, reverse=True):
        wm.delete_rows(r, 1)
    print('Metadonnees : %d lignes supprimees, %d lignes corrigees' % (len(a_supprimer), n_corrige))

    # ---- Audit ---------------------------------------------------------------
    wa = wb['Audit']
    a_supprimer_audit = []
    for r in range(2, wa.max_row + 1):
        feuille, colonne = wa.cell(r, 1).value, str(wa.cell(r, 2).value or '').strip()
        if not feuille:
            continue
        if (feuille, colonne) in RETIRES_TITRES:
            a_supprimer_audit.append(r)
        elif (feuille, colonne) in REMPLACES_TITRES:
            fic, ser, meth = REMPLACES_TITRES[(feuille, colonne)]
            wa.cell(r, 3).value = fic
            wa.cell(r, 4).value = ser
            wa.cell(r, 5).value = 'REMPLACEE (fidelite)'
            wa.cell(r, 9).value = meth
    for r in sorted(a_supprimer_audit, reverse=True):
        wa.delete_rows(r, 1)
    print('Audit : %d lignes supprimees' % len(a_supprimer_audit))

    # ---- Sommaire (recalcul complet depuis Metadonnees + comptage colonnes) --
    if 'Sommaire' in wb.sheetnames:
        del wb['Sommaire']
    ws_new = wb.create_sheet('Sommaire', 0)
    ws_new.append(['Branche', 'Colonnes', 'dont cibles', 'dont indicateurs',
                   'infra-annuels', 'alignement non vérifié'])
    from collections import defaultdict
    stats = defaultdict(lambda: [0, 0, 0, 0, 0])
    for r in wm.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        s = stats[r[0]]
        s[0] += 1
        if r[2] in ('Variable cible', 'Cible alternative'):
            s[1] += 1
        else:
            s[2] += 1
        if r[3] in ('Trimestrielle', 'Mensuelle'):
            s[3] += 1
        if r[14] == 'non vérifié':
            s[4] += 1

    branches = [f for f in wb.sheetnames
                if f not in ('Sommaire', 'Métadonnées', 'Audit', 'Ajouts',
                             'Ajouts non classés', 'Lisez-moi', 'Colonnes retirées')]
    tot = [0, 0, 0, 0, 0]
    for b in branches:
        s = stats.get(b, [0, 0, 0, 0, 0])
        ws_new.append([b] + s)
        rr = ws_new.max_row
        ws_new.cell(rr, 1).hyperlink = "#'%s'!A1" % b
        ws_new.cell(rr, 1).font = Font(color='0563C1', underline='single')
        for c in range(2, 7):
            ws_new.cell(rr, c).alignment = Alignment(horizontal='center')
        if s[4]:
            ws_new.cell(rr, 6).fill = PatternFill('solid', fgColor='FFC7CE')
        tot = [a + bb for a, bb in zip(tot, s)]
    ws_new.append(['TOTAL'] + tot)
    for c in range(1, 7):
        ws_new.cell(ws_new.max_row, c).font = Font(bold=True)
    for c, w in enumerate((32, 12, 13, 16, 15, 22), start=1):
        ws_new.column_dimensions[get_column_letter(c)].width = w
        cel = ws_new.cell(1, c)
        cel.font = F_EN
        cel.fill = EN
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_new.row_dimensions[1].height = 34
    ws_new.freeze_panes = 'A2'
    ws_new.auto_filter.ref = 'A1:F%d' % ws_new.max_row
    ws_new.sheet_view.showGridLines = False
    print('Sommaire recalcule : %d colonnes au total' % tot[0])

    wb.save(CLASSEUR)
    print('\n%s enregistre.' % CLASSEUR.name)


if __name__ == '__main__':
    main()
