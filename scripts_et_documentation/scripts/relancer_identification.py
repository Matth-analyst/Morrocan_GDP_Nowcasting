# -*- coding: utf-8 -*-
"""
Reprend les colonnes restees "source non identifiée" et celles sans fichier
source, en les confrontant a l'index COMPLET du data lake.

L'identification initiale avait ete faite avec un index incomplet : le lecteur
ne reconnaissait alors ni les dates du type "déc-07", ni le format "2014:1" des
comptes nationaux, ni les fichiers en format long (IPAI). L'index couvre
desormais ces cas, d'ou cette reprise.

Une colonne n'est reecrite que si l'appariement atteint 85 % des points.
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parent))
from catalogue_lake import charger_lake                          # noqa: E402
from completer_classeur import lire_ipai, lire_panel_annuel      # noqa: E402
from enrichir_classeur import sansacc                            # noqa: E402
import reconstruire_classeur as R                                # noqa: E402

CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'
EXCLURE = ['bdd sectoriel', 'etude_sectorielle', 'manar_panel', '_transfert']


def catalogue():
    """Toutes les series du lac, hors bases consolidees non auditables."""
    cat = []
    for s in charger_lake(ROOT) + lire_ipai():
        if any(x in sansacc(s['fichier']) for x in EXCLURE):
            continue
        cat.append({'origine': 'data lake', 'tableau': s['fichier'],
                    'serie': s['serie'], 'freq': s['freq'],
                    'valeurs': s['valeurs']})
    return cat


def main():
    wb = openpyxl.load_workbook(CLASSEUR)
    wa = wb['Audit']
    cat = catalogue()
    print('%d series candidates dans le lac (hors bases consolidees)' % len(cat))

    # colonnes a reprendre : statut NON identifiée, ou sans fichier source
    cibles = []
    for r in range(2, wa.max_row + 1):
        statut = str(wa.cell(r, 5).value or '')
        fichier = wa.cell(r, 3).value
        if statut.startswith('NON') or fichier in (None, '', '?'):
            cibles.append((r, wa.cell(r, 1).value, wa.cell(r, 2).value, statut))
    print('%d colonnes a reprendre\n' % len(cibles))

    ok = ko = 0
    jaune = PatternFill('solid', fgColor='FFF2CC')
    vert = PatternFill('solid', fgColor='C6EFCE')

    for ligne, feuille, colonne, statut in cibles:
        if feuille not in wb.sheetnames:
            continue
        ws = wb[feuille]
        # localiser la colonne dans un bloc D'ORIGINE (axe en ligne 4)
        col = axe = None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(3, c).value).strip() == str(colonne).strip():
                a = None
                for cc in range(c, 0, -1):
                    v = ws.cell(4, cc).value
                    if isinstance(v, str) and v.strip() in ('Trimestre', 'Mois'):
                        a = cc
                        break
                if a is not None:
                    col, axe = c, a
                    break
        if col is None:
            print('  %-24s %-42s colonne introuvable' % (feuille[:24], str(colonne)[:42]))
            ko += 1
            continue

        type_axe = ws.cell(4, axe).value.strip()
        obs = []
        for r in range(5, ws.max_row + 1):
            k = R.cle_axe(ws.cell(r, axe).value, type_axe)
            v = ws.cell(r, col).value
            if k and isinstance(v, (int, float)):
                obs.append((k, float(v)))
        if not obs:
            ko += 1
            continue

        s, methode, score = R.apparier(obs, cat, type_axe)
        if s is None:
            print('  %-24s %-42s toujours non identifiee (%d/%d)'
                  % (feuille[:24], str(colonne)[:42], score, len(obs)))
            ko += 1
            continue

        vals = s['valeurs']
        ecrits = 0
        for r in range(5, ws.max_row + 1):
            k = R.cle_axe(ws.cell(r, axe).value, type_axe)
            if k:
                ws.cell(r, col).value = vals.get(k)
                if k in vals:
                    ecrits += 1
        # trace dans la ligne d'audit
        wa.cell(ligne, 3).value = s['tableau']
        wa.cell(ligne, 4).value = s['serie']
        wa.cell(ligne, 5).value = ('REALIGNEE (reprise)' if methode == 'position'
                                   else 'verifiee (reprise)')
        wa.cell(ligne, 5).fill = jaune if methode == 'position' else vert
        wa.cell(ligne, 6).value = ecrits
        o = sorted(vals)
        wa.cell(ligne, 7).value, wa.cell(ligne, 8).value = o[0], o[-1]
        wa.cell(ligne, 9).value = 'identifiee lors de la reprise (index complet)'
        print('  %-24s %-42s -> %s :: %s  [%s, %d/%d]'
              % (feuille[:24], str(colonne)[:42], Path(s['tableau']).name[:34],
                 s['serie'][:28], methode, score, len(obs)))
        ok += 1

    wb.save(CLASSEUR)
    print('\n%d colonnes identifiees et reecrites, %d toujours sans source' % (ok, ko))


if __name__ == '__main__':
    main()
