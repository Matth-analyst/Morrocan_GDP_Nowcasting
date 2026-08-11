# -*- coding: utf-8 -*-
"""
Met a jour les feuilles de service apres restaurer_20_restantes.py :
- Metadonnees : ajoute une ligne par colonne restauree
- Colonnes retirées : ne garde que l'IPAI (seule colonne encore retiree)
- Sommaire : recalcule les compteurs
"""
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'

from restaurer_20_restantes import A_RESTAURER, STATUT, AUTEUR  # noqa: E402

F_EN = Font(bold=True, color='FFFFFF', size=10)
EN = PatternFill('solid', fgColor='2F5597')
ORANGE = PatternFill('solid', fgColor='FFE699')


def cle_annee(k):
    import re
    m = re.search(r'(19|20)\d{2}', str(k))
    return int(m.group(0)) if m else None


def main():
    wb = openpyxl.load_workbook(CLASSEUR)

    # ---- Metadonnees ---------------------------------------------------------
    wm = wb['Métadonnées']
    n_ajoutees = 0
    for feuille, titre, inst in A_RESTAURER:
        ws = wb[feuille]
        col, axe, type_axe = None, None, None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(3, c).value or '').strip() == titre.strip():
                col = c
                break
        if col is None:
            continue
        for cc in range(col, 0, -1):
            v3, v4 = ws.cell(3, cc).value, ws.cell(4, cc).value
            if v3 in ('Trimestre', 'Mois') or v4 in ('Trimestre', 'Mois'):
                axe, type_axe = cc, (v3 if v3 in ('Trimestre', 'Mois') else v4)
                break
        vals = []
        for r in range(5, ws.max_row + 1):
            k, v = ws.cell(r, axe).value, ws.cell(r, col).value
            if k is not None and isinstance(v, (int, float)):
                vals.append((k, v))
        if not vals:
            continue
        debut, fin = vals[0][0], vals[-1][0]
        wm.append([feuille, titre.strip(), 'Indicateur', type_axe,
                  debut, fin, cle_annee(debut), cle_annee(fin), len(vals),
                  inst, 'communiqué par l’auteur de la base', AUTEUR,
                  titre.strip(), 'Restaurée — source confirmée par l’auteur', STATUT])
        wm.cell(wm.max_row, 15).fill = ORANGE
        n_ajoutees += 1
    print('Metadonnees : %d colonnes restaurees ajoutees' % n_ajoutees)

    # ---- Colonnes retirées : ne garde que l'IPAI ------------------------------
    if 'Colonnes retirées' in wb.sheetnames:
        wr = wb['Colonnes retirées']
        a_effacer = []
        titres_restaures = {t.strip() for _, t, _ in A_RESTAURER}
        for r in range(5, wr.max_row + 1):
            titre = str(wr.cell(r, 2).value or '').strip()
            if titre in titres_restaures:
                a_effacer.append(r)
        for r in sorted(a_effacer, reverse=True):
            wr.delete_rows(r, 1)
        wr.cell(2, 1).value = (
            'Seule l’IPAI en niveau reste retiree : aucune source primaire '
            'trouvee, et contrairement aux autres, aucune confirmation '
            'institutionnelle n’a ete demandee pour celle-ci (un indice '
            'reconstruit par chainage des variations trimestrielles est '
            'disponible dans la feuille Immobilier a la place).')
        print('Colonnes retirées : %d ligne(s) restante(s)' % (wr.max_row - 4))

    # ---- Sommaire --------------------------------------------------------------
    if 'Sommaire' in wb.sheetnames:
        del wb['Sommaire']
    ws_new = wb.create_sheet('Sommaire', 0)
    ws_new.append(['Branche', 'Colonnes', 'dont cibles', 'dont indicateurs',
                   'infra-annuels', 'alignement non vérifié'])
    stats = defaultdict(lambda: [0, 0, 0, 0, 0])
    for r in wm.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        s = stats[r[0]]
        s[0] += 1
        if r[2] in ('Variable cible', 'Cible alternative'):
            s[1] += 1
        else:
            s[2] += 1
        if r[3] in ('Trimestrielle', 'Mensuelle'):
            s[3] += 1
        if r[14] in ('non vérifié', STATUT):
            s[4] += 1

    branches = [f for f in wb.sheetnames
                if f not in ('Sommaire', 'Métadonnées', 'Audit', 'Ajouts',
                             'Ajouts non classés', 'Lisez-moi', 'Colonnes retirées')]
    tot = [0, 0, 0, 0, 0]
    for b in branches:
        s = stats.get(b, [0, 0, 0, 0, 0])
        ws_new.append([b] + s)
        rr = ws_new.max_row
        ws_new.cell(rr, 1).hyperlink = "#'%s'!A1" % b
        ws_new.cell(rr, 1).font = Font(color='0563C1', underline='single')
        for c in range(2, 7):
            ws_new.cell(rr, c).alignment = Alignment(horizontal='center')
        if s[4]:
            ws_new.cell(rr, 6).fill = PatternFill('solid', fgColor='FFEB9C')
        tot = [a + bb for a, bb in zip(tot, s)]
    ws_new.append(['TOTAL'] + tot)
    for c in range(1, 7):
        ws_new.cell(ws_new.max_row, c).font = Font(bold=True)
    for c, w in enumerate((32, 12, 13, 16, 15, 22), start=1):
        ws_new.column_dimensions[get_column_letter(c)].width = w
        cel = ws_new.cell(1, c)
        cel.font = F_EN
        cel.fill = EN
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_new.row_dimensions[1].height = 34
    ws_new.freeze_panes = 'A2'
    ws_new.auto_filter.ref = 'A1:F%d' % ws_new.max_row
    ws_new.sheet_view.showGridLines = False
    print('Sommaire recalcule : %d colonnes au total' % tot[0])

    wb.save(CLASSEUR)
    print('\n%s enregistre.' % CLASSEUR.name)


if __name__ == '__main__':
    main()
