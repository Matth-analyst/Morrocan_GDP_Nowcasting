# -*- coding: utf-8 -*-
"""
Reconstruit un indice IPAI (niveau, categorie Global) par chainage des
variations trimestrielles extraites des 72 bulletins Bank Al-Maghrib/ANCFCC.

Methode
-------
Index[premier trimestre disponible] = 100 (base interne, PAS la base
officielle 100 = 2006 de Bank Al-Maghrib, que nous n'avons pas).
Pour chaque trimestre suivant CONSECUTIF (sans trou) :
    Index[t] = Index[t-1] * (1 + variation_trimestrielle[t] / 100)

Un trou dans les trimestres casse la chaine : le chainage repart de 100 au
premier trimestre disponible apres le trou, et un nouveau "segment" commence.
On ne comble jamais un trou par interpolation : mieux vaut deux segments
honnetes qu'une valeur inventee.

Sortie : colonne ajoutee dans la feuille Immobilier, avec la formule et les
limites documentees en clair sur la colonne elle-meme.
"""
import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'
IPAI = ROOT / 'Bank Al-Maghrib et ANCFCC' / 'immobilier_ipai' / 'ipai_variations_propre.csv'


def cle_chrono(trimestre):
    """'T3-2014' -> (2014, 3), pour un tri CHRONOLOGIQUE (le tri alphabetique
    de la chaine est faux : 'T1-2023' < 'T4-2022' alphabetiquement alors que
    T4-2022 est anterieur)."""
    t, a = trimestre.split('-')
    return (int(a), int(t[1]))


def suivant(cle):
    a, t = cle
    return (a, t + 1) if t < 4 else (a + 1, 1)


def main():
    rows = list(csv.DictReader(open(IPAI, encoding='utf-8-sig'), delimiter=';'))
    glob = [r for r in rows if r['categorie'] == 'Global'
            and r['indicateur'].strip().startswith('prix')
            and r['variation_trimestrielle'].strip()]
    par_trim = {}
    for r in glob:
        k = cle_chrono(r['trimestre'])
        v = float(r['variation_trimestrielle'].replace(',', '.'))
        # en cas de doublon (deux bulletins pour le meme trimestre), on garde
        # la lecture la plus fiable
        rang = {'haute (lecture manuelle verifiee)': 3, 'haute': 2,
               'basse': 1, 'approximation (texte source qualitatif uniquement, sans chiffre precis)': 0}
        if k not in par_trim or rang.get(r['confiance'], 0) > par_trim[k][1]:
            par_trim[k] = (v, rang.get(r['confiance'], 0))
    variations = {k: v[0] for k, v in par_trim.items()}

    trimestres = sorted(variations)
    print('Trimestres disponibles (%d), triés chronologiquement :' % len(trimestres))
    print('  de T%d-%d à T%d-%d' % (trimestres[0][1], trimestres[0][0],
                                    trimestres[-1][1], trimestres[-1][0]))

    # decoupage en segments consecutifs
    segments = [[trimestres[0]]]
    for k in trimestres[1:]:
        if k == suivant(segments[-1][-1]):
            segments[-1].append(k)
        else:
            segments.append([k])
    print('\n%d segment(s) continu(s) (un trou casse la chaine) :' % len(segments))
    for seg in segments:
        print('  T%d-%d -> T%d-%d  (%d trimestres)'
              % (seg[0][1], seg[0][0], seg[-1][1], seg[-1][0], len(seg)))

    # chainage
    index = {}
    for seg in segments:
        index[seg[0]] = 100.0
        for i in range(1, len(seg)):
            k, kp = seg[i], seg[i - 1]
            index[k] = round(index[kp] * (1 + variations[k] / 100), 4)

    plus_long = max(segments, key=len)
    print('\nSegment retenu pour la colonne (le plus long, %d trimestres) : '
          'T%d-%d -> T%d-%d' % (len(plus_long), plus_long[0][1], plus_long[0][0],
                                plus_long[-1][1], plus_long[-1][0]))
    for k in plus_long[:3] + plus_long[-3:]:
        print('   T%d-%d : %.2f' % (k[1], k[0], index[k]))

    # ---- ecriture dans le classeur ------------------------------------------
    wb = openpyxl.load_workbook(CLASSEUR)
    ws = wb['Immobilier']

    axe_trim = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(3, c).value == 'Trimestre' or ws.cell(4, c).value == 'Trimestre':
            axe_trim = c

    # supprime une precedente version de cette colonne, pour eviter tout doublon
    # en cas de reexecution du script
    for c in range(ws.max_column, 0, -1):
        if isinstance(ws.cell(3, c).value, str) and ws.cell(3, c).value.startswith(
                'IPAI Global (indice reconstruit'):
            ws.delete_cols(c, 1)

    axe_trim = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(3, c).value == 'Trimestre' or ws.cell(4, c).value == 'Trimestre':
            axe_trim = c
    if axe_trim is None:
        sys.exit('axe Trimestre introuvable dans Immobilier')

    c_ins = axe_trim + 1
    while not all(ws.cell(r, c_ins).value in (None, '') for r in range(1, 12)):
        c_ins += 1
    ws.insert_cols(c_ins, 1)

    ws.cell(3, c_ins).value = 'IPAI Global (indice reconstruit, base 100 interne)'
    ws.cell(4, c_ins).value = ('Bank Al-Maghrib/ANCFCC (chaînage des variations '
                               'trimestrielles extraites de 72 bulletins) — '
                               'NON recalé sur la base officielle 100=2006')
    ws.cell(2, c_ins).value = (
        'RECONSTRUCTION : base 100 = T%d-%d (premier trimestre du plus long '
        'segment continu). Chaque valeur = precedente x (1 + variation '
        'trimestrielle du bulletin). Ne couvre PAS 2006-2009 ni la periode '
        'recente (derniere variation disponible : T%d-%d). Ne pas comparer '
        'en niveau a un indice officiel base 2006 ; utilisable pour la '
        'DYNAMIQUE (croissance) sur la periode couverte.'
        % (plus_long[0][1], plus_long[0][0], plus_long[-1][1], plus_long[-1][0]))
    ws.cell(3, c_ins).fill = PatternFill('solid', fgColor='FFE699')
    ws.cell(3, c_ins).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(2, c_ins).alignment = Alignment(wrap_text=True, vertical='top')

    lignes_axe = {}
    for r in range(5, ws.max_row + 1):
        v = ws.cell(r, axe_trim).value
        if isinstance(v, str) and v.startswith('T'):
            lignes_axe[cle_chrono(v)] = r
    derniere = max(lignes_axe.values()) if lignes_axe else 4
    for k in plus_long:
        if k not in lignes_axe:
            derniere += 1
            ws.cell(derniere, axe_trim).value = 'T%d-%d' % (k[1], k[0])
            lignes_axe[k] = derniere
        ws.cell(lignes_axe[k], c_ins).value = index[k]

    ws.row_dimensions[2].height = 60
    ws.row_dimensions[3].height = 46

    wb.save(CLASSEUR)
    print('\n%s : colonne "IPAI Global (indice reconstruit...)" ajoutee a Immobilier'
          % CLASSEUR.name)
    return plus_long, index


if __name__ == '__main__':
    main()
