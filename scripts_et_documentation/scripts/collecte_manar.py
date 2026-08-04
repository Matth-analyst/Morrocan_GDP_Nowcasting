# -*- coding: utf-8 -*-
"""
Collecte des tableaux Manar-Stat (DEPF / Ministere de l'Economie et des Finances, Maroc)
Domaine Sectoriel, a partir de la rubrique "Mines" jusqu'a la fin de l'arborescence.

Mecanisme (aucun navigateur requis) :
  1. DWR  : AjaxConsultation.getListDomaineTableau -> arborescence complete
  2. POST : Consultation_consulterTable (hiddenDomaine, hiddenTableau) -> page ZK
  3. ZK AU: onClick sur le bouton exportToXLS -> URL de media
  4. GET  : telechargement du .xlsx, puis conversion en .csv

Usage :
    python collecte_manar.py --rubrique mines
    python collecte_manar.py --rubrique all
    python collecte_manar.py --list
"""

import argparse
import io
import json
import os
import re
import sys
import time
import unicodedata
from datetime import datetime

import openpyxl
import requests

BASE = "https://manar.finances.gov.ma/manar"
ROOT = os.path.dirname(os.path.abspath(__file__))
DELAI = 8.0  # secondes entre deux tableaux (le portail sature facilement)
UA = ("Mozilla/5.0 (compatible; recherche academique nowcasting PIB Maroc; "
      "contact komlanadjohmatthania@gmail.com)")

AN_MIN = 1960   # borne basse demandee au filtre de periodes
AN_MAX = 2027   # borne haute demandee au filtre de periodes

ID_SECTORIEL = 432
ID_SECONDAIRE = 718
ID_TERTIAIRE = 719
ID_MINES = 285

# rubrique de 1er niveau (id) -> (dossier parent, sous-dossier)
RUBRIQUES = {
    243:  ("manar_primaire", "cultures"),
    244:  ("manar_primaire", "elevage"),
    245:  ("manar_primaire", "peche"),
    246:  ("manar_primaire", "climatologie"),
    285:  ("manar_secondaire", "mines"),
    435:  ("manar_secondaire", "construction"),
    364:  ("manar_secondaire", "eau"),
    434:  ("manar_secondaire", "energie"),
    440:  ("manar_secondaire", "industrie"),
    439:  ("manar_tertiaire", "transports"),
    438:  ("manar_tertiaire", "tourisme"),
    436:  ("manar_tertiaire", "telecommunications"),
    437:  ("manar_tertiaire", "assurances"),
}

# ordre de traitement : Primaire, Secondaire, puis Tertiaire
ORDRE = [243, 244, 245, 246, 285, 364, 434, 435, 440, 439, 438, 436, 437]


# --------------------------------------------------------------------------
# Parseur de litteral JavaScript (la reponse DWR n'est pas du JSON valide :
# cles non quotees, references cycliques, valeurs null/true/false)
# --------------------------------------------------------------------------
class JsLiteral:
    def __init__(self, txt):
        self.s = txt
        self.i = 0

    def ws(self):
        while self.i < len(self.s) and self.s[self.i] in " \t\r\n":
            self.i += 1

    def parse(self):
        self.ws()
        c = self.s[self.i]
        if c == "{":
            return self.obj()
        if c == "[":
            return self.arr()
        if c in "\"'":
            return self.string()
        return self.literal()

    def obj(self):
        self.i += 1  # {
        out = {}
        self.ws()
        if self.s[self.i] == "}":
            self.i += 1
            return out
        while True:
            self.ws()
            if self.s[self.i] in "\"'":
                key = self.string()
            else:
                j = self.i
                while self.s[self.i] not in ":":
                    self.i += 1
                key = self.s[j:self.i].strip()
            self.ws()
            self.i += 1  # :
            out[key] = self.parse()
            self.ws()
            if self.s[self.i] == ",":
                self.i += 1
                continue
            if self.s[self.i] == "}":
                self.i += 1
                return out

    def arr(self):
        self.i += 1  # [
        out = []
        self.ws()
        if self.s[self.i] == "]":
            self.i += 1
            return out
        while True:
            out.append(self.parse())
            self.ws()
            if self.s[self.i] == ",":
                self.i += 1
                continue
            if self.s[self.i] == "]":
                self.i += 1
                return out

    def string(self):
        q = self.s[self.i]
        self.i += 1
        buf = []
        while True:
            c = self.s[self.i]
            if c == "\\":
                nxt = self.s[self.i + 1]
                if nxt == "u":
                    buf.append(chr(int(self.s[self.i + 2:self.i + 6], 16)))
                    self.i += 6
                else:
                    buf.append({"n": "\n", "t": "\t", "r": "\r"}.get(nxt, nxt))
                    self.i += 2
                continue
            if c == q:
                self.i += 1
                return "".join(buf)
            buf.append(c)
            self.i += 1

    def literal(self):
        j = self.i
        while self.i < len(self.s) and self.s[self.i] not in ",}] \t\r\n":
            self.i += 1
        tok = self.s[j:self.i]
        if tok == "null":
            return None
        if tok == "true":
            return True
        if tok == "false":
            return False
        try:
            return int(tok)
        except ValueError:
            try:
                return float(tok)
            except ValueError:
                return tok  # reference DWR (s0, s1, ...)


# --------------------------------------------------------------------------
# Construction de la liste de periodes attendue par le filtre du portail.
# Formats repris a l'identique de SerieCalendrierTableau.jsp (fonction
# recuperer_date) : le serveur refuse toute autre notation.
# --------------------------------------------------------------------------
MOIS_CUM = ["Janv", "Fev", "Mars", "Avr", "Mai", "Juin",
            "Juil", "Aout", "Sep", "Oct", "Nov", "Dec"]


def periodes(frequence, an_min=AN_MIN, an_max=AN_MAX):
    f = (frequence or "").lower()
    ans = range(an_min, an_max + 1)

    if "trimestrielle cumul" in f:
        out = []
        for a in ans:
            out.append("Trim1-%d" % a)
            for i in (2, 3, 4):
                out.append("Trim1:Trim%d-%d" % (i, a))
        return out
    if "mensuelle cumul" in f:
        out = []
        for a in ans:
            out.append("%s-%d" % (MOIS_CUM[0], a))
            for j in range(1, 12):
                out.append("%s:%s-%d" % (MOIS_CUM[0], MOIS_CUM[j], a))
        return out
    if "semestrielle cumul" in f:
        out = []
        for a in ans:
            out.append("Semestre1-%d" % a)
            out.append("Semestre1:Semestre2-%d" % a)
        return out
    if "trimestrielle" in f:
        return ["%dT%d" % (a, i) for a in ans for i in (1, 2, 3, 4)]
    if "semestrielle" in f:
        return ["Semestre%d-%d" % (i, a) for a in ans for i in (1, 2)]
    if "mensuelle" in f:
        return ["%dM%02d" % (a, m) for a in ans for m in range(1, 13)]
    if "cheval" in f or "campagne" in f:
        # campagnes agricoles : 80/81 ... 98/99, 99/2000, 2000/2001 ...
        out = []
        for a in ans:
            if a == 1999:
                out.append("99/2000")
            elif a > 1999:
                out.append("%d/%d" % (a, a + 1))
            else:
                d = a % 100
                out.append("%d/%d" % (d, d + 1))
        return out
    if "annuelle" in f or f == "":
        return [str(a) for a in ans]
    return []  # frequence non geree (journaliere, bimensuelle, budgetaire...)


FREQUENCES = ["Trimestrielle cumulée", "Mensuelle cumulée", "Semestrielle cumulée",
              "Trimestrielle", "Semestrielle", "Mensuelle", "Bimensuelle",
              "Journalière", "Annuelle budgétaire", "Année à  cheval", "Annuelle"]


def detecter_frequence(html):
    """Deduit la frequence des libelles de periodes reellement affiches.

    Le mot-cle "Annuelle" apparait dans la page meme pour des tableaux
    trimestriels ou en campagne agricole : seuls les en-tetes de colonnes
    sont fiables (ils sont dans le pool de chaines du composant ZK).
    """
    m = re.search(r"'s':\[(.{0,4000})", html, re.S)
    if m:
        jetons = re.findall(r"'([^']{1,24})'", m.group(1))
        for j in jetons:
            if re.fullmatch(r"\d{4}T[1-4]", j):
                return "Trimestrielle"
            if re.fullmatch(r"\d{4}M\d{2}", j):
                return "Mensuelle"
            if re.fullmatch(r"\d{2,4}/\d{2,4}", j):
                return "Année à cheval"
            if j.startswith("Trim1:") or j.startswith("Trim1-"):
                return "Trimestrielle cumulée"
            if j.startswith("Semestre"):
                return "Semestrielle"
            if re.match(r"(Janv|Fev|Mars|Avr|Mai|Juin|Juil|Aout|Sep|Oct|Nov|Dec)[-:]", j):
                return "Mensuelle cumulée"
            if re.fullmatch(r"(19|20)\d{2}", j):
                return "Annuelle"
    for f in FREQUENCES:
        if ("'%s'" % f) in html or ('"%s"' % f) in html:
            return f
    return "Annuelle"


# --------------------------------------------------------------------------
# Client Manar
# --------------------------------------------------------------------------
class Manar:
    def __init__(self):
        self.s = requests.Session()
        self.s.headers["User-Agent"] = UA
        self.s.get(BASE + "/Consultation_domainetableau", timeout=60)
        self.jid = self.s.cookies.get("JSESSIONID")

    def arborescence(self):
        body = {
            "callCount": "1", "windowName": "",
            "c0-scriptName": "AjaxConsultation",
            "c0-methodName": "getListDomaineTableau",
            "c0-id": "0", "batchId": "0", "instanceId": "0",
            "page": "/manar/Consultation_domainetableau",
            "scriptSessionId": "aaaaaaaaaaaaaaaa/" + self.jid,
            "httpSessionId": self.jid,
        }
        r = self.s.post(
            BASE + "/dwr/call/plaincall/AjaxConsultation.getListDomaineTableau.dwr",
            data="\n".join(k + "=" + v for k, v in body.items()),
            headers={"Content-Type": "text/plain"}, timeout=180)
        r.raise_for_status()
        m = re.search(r'handleCallback\("0","0",(\[.*)\);\s*$', r.text, re.S)
        if not m:
            raise RuntimeError("reponse DWR inattendue")
        return JsLiteral(m.group(1)).parse()

    def telecharger_xlsx(self, id_domaine, id_tableau, elargir=True):
        """Ouvre le tableau, elargit la periode a tout l'historique, exporte.

        Renvoie (contenu_xlsx_bytes, frequence, filtre_applique)."""
        r = self.s.post(BASE + "/Consultation_consulterTable",
                        data={"hiddenDomaine": str(id_domaine),
                              "hiddenTableau": str(id_tableau)}, timeout=90)
        r.raise_for_status()
        html = r.text
        m_dt = re.search(r"dt:'([^']+)'", html)
        m_bt = re.search(r"'zul\.wgt\.Toolbarbutton','([^']+)',\{id:'exportToXLS'", html)
        if not (m_dt and m_bt):
            raise RuntimeError("page ZK sans bouton d'export (tableau vide ?)")
        dtid = m_dt.group(1)

        sid = [0]

        def au(data):
            sid[0] += 1
            return self.s.post(
                BASE + "/zkau", data=data,
                headers={"ZK-SID": str(sid[0]),
                         "X-Requested-With": "XMLHttpRequest",
                         "Content-Type":
                             "application/x-www-form-urlencoded;charset=UTF-8",
                         "Referer": BASE + "/Consultation_consulterTable"},
                timeout=240)

        # 1. elargissement de la periode affichee (sinon ~10 periodes seulement)
        freq = detecter_frequence(html)
        liste = periodes(freq)
        filtre = False
        m_lab = re.search(r"'zul\.wgt\.Label','([^']+)',\{id:'lab'", html)
        m_val = re.search(r"'zul\.wgt\.Button','([^']+)',\{id:'valider1'", html)
        if elargir and liste and m_lab and m_val:
            # NB : pas de virgule initiale, le portail la retire (axtemps.substring(1))
            val = json.dumps({"": ["value", ",".join(liste)]})
            rep = au({"dtid": dtid,
                      "cmd_0": "setAttr", "uuid_0": m_lab.group(1), "data_0": val,
                      "cmd_1": "setAttr", "uuid_1": m_lab.group(1), "data_1": val,
                      "cmd_2": "onClick", "uuid_2": m_val.group(1), "data_2": "{}"})
            filtre = "Merci de v" not in rep.text

        # 2. export Excel
        a = au({"dtid": dtid, "cmd_0": "onClick", "uuid_0": m_bt.group(1),
                "data_0": json.dumps({"pageX": 10, "pageY": 10, "which": 1,
                                      "x": 1, "y": 1}),
                "opt_0": "i"})
        m_url = re.search(r'"download",\["([^"]+)"\]', a.text)
        if not m_url:
            raise RuntimeError("export refuse par le serveur")
        url = "https://manar.finances.gov.ma" + m_url.group(1).replace("\\/", "/")
        f = self.s.get(url, timeout=240)
        f.raise_for_status()
        if not f.content.startswith(b"PK"):
            raise RuntimeError("fichier telecharge non valide (%d octets)"
                               % len(f.content))
        return f.content, freq, filtre


# --------------------------------------------------------------------------
# Outils
# --------------------------------------------------------------------------
def nom_fichier(libelle):
    """Nom de fichier sur : accents conserves, caracteres interdits retires."""
    n = libelle.strip()
    n = re.sub(r'[\\/:*?"<>|\r\n\t]', " ", n)
    n = re.sub(r"\s+", " ", n).strip(" .")
    return n[:150]


def aplatir(noeud, chemin=""):
    """Rend la liste des tableaux (dict) sous un noeud, en profondeur."""
    p = (chemin + " / " + noeud["libelle"]) if chemin else noeud["libelle"]
    out = []
    for t in (noeud.get("tableaux") or []):
        out.append({"id": t["id"], "libelle": (t.get("nom") or t.get("libelle")),
                    "id_domaine": noeud["id"], "chemin": p})
    for c in (noeud.get("children") or []):
        if isinstance(c, dict):
            out.extend(aplatir(c, p))
    return out


def trouver(noeuds, cible):
    for n in noeuds:
        if not isinstance(n, dict):
            continue
        if n.get("id") == cible:
            return n
        r = trouver(n.get("children") or [], cible)
        if r:
            return r
    return None


def xlsx_verifier(contenu):
    """Leve une exception si l'export ne contient aucune donnee exploitable."""
    wb = openpyxl.load_workbook(io.BytesIO(contenu), data_only=True)
    ws = wb["Données"] if "Données" in wb.sheetnames else wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        if any(c is not None and str(c).strip() for c in row):
            return
    raise RuntimeError("feuille de donnees vide")


def xlsx_vers_csv(contenu, chemin_csv):
    """Ecrit le CSV et renvoie (premiere_periode, derniere_periode, nb_lignes)."""
    wb = openpyxl.load_workbook(io.BytesIO(contenu), data_only=True)
    ws = wb["Données"] if "Données" in wb.sheetnames else wb.worksheets[0]
    lignes = []
    for row in ws.iter_rows(values_only=True):
        vals = ["" if v is None else str(v) for v in row]
        while vals and vals[-1] == "":
            vals.pop()
        if any(v.strip() for v in vals):
            lignes.append(vals)
    if not lignes:
        raise RuntimeError("feuille de donnees vide")

    largeur = max(len(l) for l in lignes)
    lignes = [l + [""] * (largeur - len(l)) for l in lignes]

    # On demande volontairement une grille de periodes tres large : on retire
    # ici les colonnes sans aucune donnee pour ne garder que l'etendue reelle.
    garder = [0] + [j for j in range(1, largeur)
                    if lignes[0][j].strip()
                    and any(l[j].strip() for l in lignes[1:])]
    if len(garder) < 2:
        raise RuntimeError("aucune periode renseignee")
    lignes = [[l[j] for j in garder] for l in lignes]

    import csv
    with open(chemin_csv, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerows(lignes)

    entete = lignes[0][1:]
    return entete[0], entete[-1], len(lignes) - 1


# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rubrique", default="mines",
                    help="mines|eau|energie|construction|industrie|transports|"
                         "tourisme|telecommunications|assurances|all")
    ap.add_argument("--list", action="store_true",
                    help="afficher l'inventaire sans rien telecharger")
    ap.add_argument("--reprise", action="store_true",
                    help="ne traiter que les tableaux dont le CSV manque")
    args = ap.parse_args()

    print("Connexion a Manar-Stat ...")
    for essai in range(5):
        try:
            m = Manar()
            break
        except Exception as err:
            print("  connexion refusee (%s), nouvelle tentative dans 60 s"
                  % str(err)[:60])
            time.sleep(60)
    else:
        sys.exit("portail injoignable — reessayer plus tard")
    arbre = m.arborescence()
    sect = trouver(arbre, ID_SECTORIEL)
    if sect is None:
        sys.exit("domaine Sectoriel introuvable")

    par_rubrique = {}
    for rid in ORDRE:
        noeud = trouver([sect], rid)
        if noeud is None:
            print("  !! rubrique %s introuvable" % rid)
            continue
        par_rubrique[rid] = aplatir(noeud)

    if args.list:
        for rid in ORDRE:
            dossier = "/".join(RUBRIQUES[rid])
            tabs = par_rubrique.get(rid, [])
            print("\n=== %s  (%d tableaux) -> %s" %
                  (trouver([sect], rid)["libelle"], len(tabs), dossier))
            for t in tabs:
                print("   %6d  %s" % (t["id"], t["libelle"]))
        return

    if args.rubrique == "all":
        cibles = ORDRE
    else:
        inv = {v[1]: k for k, v in RUBRIQUES.items()}
        if args.rubrique not in inv:
            sys.exit("rubrique inconnue : " + args.rubrique)
        cibles = [inv[args.rubrique]]

    journal = []
    for rid in cibles:
        parent, sous = RUBRIQUES[rid]
        dossier = os.path.join(ROOT, parent, sous)
        os.makedirs(dossier, exist_ok=True)
        tabs = par_rubrique.get(rid, [])
        if args.reprise:
            tabs = [t for t in tabs
                    if not os.path.exists(os.path.join(
                        dossier, nom_fichier(t["libelle"]) + ".csv"))]
            if not tabs:
                continue
        print("\n=== %s : %d tableaux -> %s/%s" %
              (trouver([sect], rid)["libelle"], len(tabs), parent, sous))

        for k, t in enumerate(tabs, 1):
            base = nom_fichier(t["libelle"])
            entree = {"rubrique": "%s/%s" % (parent, sous), "id": t["id"],
                      "libelle": t["libelle"], "chemin": t["chemin"]}
            try:
                contenu = freq = filtre = None
                derniere = None
                # 3 tentatives : grille elargie, puis export par defaut, chaque
                # fois avec une session neuve (le serveur coupe la connexion sur
                # les gros tableaux et la session ne s'en remet pas)
                csv_tmp = os.path.join(dossier, base + ".csv")
                res = None
                for tentative, elargir in ((1, True), (2, True), (3, False)):
                    try:
                        contenu, freq, filtre = m.telecharger_xlsx(
                            t["id_domaine"], t["id"], elargir=elargir)
                        xlsx_verifier(contenu)
                        # la conversion fait partie de la tentative : si la
                        # grille elargie ressort sans periode exploitable, on
                        # veut retomber sur l'export par defaut
                        res = xlsx_vers_csv(contenu, csv_tmp)
                        break
                    except Exception as err:
                        derniere = err
                        contenu = None
                        if tentative < 3:
                            # temporisation croissante : le portail coupe les
                            # connexions quand on le sollicite trop vite
                            time.sleep(30 * tentative)
                            try:
                                m = Manar()
                            except Exception:
                                time.sleep(60)
                                try:
                                    m = Manar()
                                except Exception:
                                    pass
                if res is None:
                    raise derniere
                with open(os.path.join(dossier, base + ".xlsx"), "wb") as fh:
                    fh.write(contenu)
                p1, p2, n = res
                entree.update(statut="ok", debut=p1, fin=p2, lignes=n,
                              frequence=freq, filtre=filtre)
                print("  [%2d/%2d] OK   %-58.58s %-14s %s -> %s (%d series)%s"
                      % (k, len(tabs), base, freq, p1, p2, n,
                         "" if filtre else "  [periode non elargie]"))
            except Exception as e:
                entree.update(statut="echec", raison=str(e)[:200])
                print("  [%2d/%2d] ECHEC %-68.68s  %s"
                      % (k, len(tabs), base, str(e)[:80]))
            journal.append(entree)
            time.sleep(DELAI)

    ecrire_journal(journal, cibles, sect)


def ecrire_journal(journal, cibles, sect):
    chemin = os.path.join(ROOT, "log_collecte.md")
    deja = ""
    if os.path.exists(chemin):
        deja = open(chemin, encoding="utf-8").read()
    else:
        deja = ("# Journal de collecte Manar-Stat\n\n"
                "Source : Banque de donnees Manar-Stat, DEPF, Ministere de l'Economie\n"
                "et des Finances du Maroc — https://manar.finances.gov.ma\n\n"
                "Methode : export XLSX natif du portail (bouton \"Exporter vers Excel\"),\n"
                "converti en CSV (separateur `;`, encodage UTF-8 BOM). Delai de %.0f s\n"
                "entre chaque tableau.\n" % DELAI)

    ok = [e for e in journal if e["statut"] == "ok"]
    ko = [e for e in journal if e["statut"] != "ok"]
    lignes = ["\n\n---\n\n## Execution du %s\n"
              % datetime.now().strftime("%d/%m/%Y a %H:%M"),
              "%d tableaux traites — %d recuperes, %d en echec.\n"
              % (len(journal), len(ok), len(ko))]

    if ok:
        lignes.append("\n### Tableaux recuperes\n")
        lignes.append("| Rubrique | Id | Tableau | Frequence | Debut | Fin | Series | Periode elargie |")
        lignes.append("|---|---|---|---|---|---|---|---|")
        for e in ok:
            lignes.append("| %s | %d | %s | %s | %s | %s | %d | %s |"
                          % (e["rubrique"], e["id"], e["libelle"].replace("|", "/"),
                             e.get("frequence", "?"), e["debut"], e["fin"],
                             e["lignes"], "oui" if e.get("filtre") else "non"))
    if ko:
        lignes.append("\n### Echecs\n")
        lignes.append("| Rubrique | Id | Tableau | Raison |")
        lignes.append("|---|---|---|---|")
        for e in ko:
            lignes.append("| %s | %d | %s | %s |"
                          % (e["rubrique"], e["id"], e["libelle"].replace("|", "/"),
                             e["raison"].replace("|", "/")))

    with open(chemin, "w", encoding="utf-8") as fh:
        fh.write(deja + "\n".join(lignes) + "\n")
    print("\nJournal mis a jour : log_collecte.md")


if __name__ == "__main__":
    main()
