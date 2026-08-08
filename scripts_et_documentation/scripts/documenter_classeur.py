# -*- coding: utf-8 -*-
"""
Met a jour les feuilles de service de Etude_sectorielle_Maroc_2_complete.xlsx :

* Metadonnees : ajout du FICHIER SOURCE et du nom exact de la serie dans ce
  fichier, pour chaque ligne existante, puis ajout d'une ligne par serie
  ajoutee depuis le data lake.
* Sommaire    : recapitulatif par branche (colonnes d'origine + ajouts).
* Audit       : inchangee (elle documente la reconstruction de _2).

Le fichier est modifie sur place.
"""

import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'

EN_TETE = PatternFill('solid', fgColor='2F5597')
F_EN = Font(bold=True, color='FFFFFF', size=10)
F_TITRE = Font(bold=True, size=14, color='1F3864')
F_SOUS = Font(italic=True, size=10, color='595959')
CENTRE = Alignment(horizontal='center', vertical='center', wrap_text=True)


def norm(s):
    s = unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]', '', s.lower())


def main():
    wb = openpyxl.load_workbook(CLASSEUR)
    wm = wb['Métadonnées']
    wa = wb['Audit']
    waj = wb['Ajouts']

    # ---- index des sources connues, depuis Audit ---------------------------
    src_audit = {}
    for r in wa.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        feuille, colonne, fichier, serie, statut = r[0], r[1], r[2], r[3], r[4]
        src_audit[(norm(feuille), norm(colonne)[:34])] = (fichier, serie, statut)

    # ---- 1. enrichissement des lignes existantes ---------------------------
    ncol = wm.max_column
    col_fic, col_ser, col_org = ncol + 1, ncol + 2, ncol + 3
    for c, titre in ((col_fic, 'Fichier source'),
                     (col_ser, 'Série dans le fichier'),
                     (col_org, 'Origine')):
        cel = wm.cell(1, c)
        cel.value = titre
        cel.font = F_EN
        cel.fill = EN_TETE
        cel.alignment = CENTRE

    rempli = vide = 0
    for r in range(2, wm.max_row + 1):
        branche, serie = wm.cell(r, 1).value, wm.cell(r, 2).value
        if not branche or not serie:
            continue
        cle = (norm(branche), norm(serie)[:34])
        info = src_audit.get(cle)
        if info is None:
            for (b, s), v in src_audit.items():
                if b == norm(branche) and (s.startswith(norm(serie)[:20])
                                           or norm(serie).startswith(s[:20])):
                    info = v
                    break
        if info:
            fichier, serie_src, statut = info
            wm.cell(r, col_fic).value = fichier
            wm.cell(r, col_ser).value = serie_src
            wm.cell(r, col_org).value = ('source non identifiée'
                                         if str(statut).startswith('NON')
                                         else 'colonne d’origine')
            rempli += 1
        else:
            wm.cell(r, col_fic).value = '—'
            wm.cell(r, col_org).value = 'colonne d’origine'
            vide += 1
    print('Metadonnees : %d lignes documentees, %d sans source retrouvee'
          % (rempli, vide))

    # ---- 2. ajout des series venues du data lake ---------------------------
    n_aj = 0
    for r in waj.iter_rows(min_row=5, values_only=True):
        if not r or not r[0]:
            continue
        branche, serie, freq, fichier, deb, fin, nobs = r[:7]
        wm.append([branche, serie, Path(str(fichier)).parts[0] if fichier else '',
                   freq, deb, fin, nobs, fichier, serie, 'ajout data lake'])
        n_aj += 1
    print('Metadonnees : %d lignes ajoutees pour les series du data lake' % n_aj)

    largeurs = (24, 48, 30, 14, 12, 12, 10, 56, 44, 22)
    for i, w in enumerate(largeurs, start=1):
        wm.column_dimensions[get_column_letter(i)].width = w
    for c in range(1, wm.max_column + 1):
        cel = wm.cell(1, c)
        cel.font = F_EN
        cel.fill = EN_TETE
        cel.alignment = CENTRE
    wm.row_dimensions[1].height = 30
    wm.freeze_panes = 'A2'
    wm.auto_filter.ref = 'A1:%s%d' % (get_column_letter(wm.max_column), wm.max_row)

    # ---- 3. Sommaire -------------------------------------------------------
    if 'Sommaire' in wb.sheetnames:
        del wb['Sommaire']
    ws = wb.create_sheet('Sommaire', 0)
    ws.append(['ÉTUDE SECTORIELLE MAROC — tableau de synthèse complété'])
    ws.cell(1, 1).font = Font(bold=True, size=16, color='1F3864')
    ws.append(['Colonnes d’origine (reconstruites et vérifiées) + séries '
               'ajoutées depuis le data lake.'])
    ws.cell(2, 1).font = F_SOUS
    ws.append(['Provenance de chaque colonne : feuilles Métadonnées (toutes), '
               'Audit (colonnes d’origine), Ajouts (nouvelles séries).'])
    ws.cell(3, 1).font = F_SOUS
    ws.append([])
    ws.append(['Branche', 'Colonnes d’origine', 'dont source vérifiée',
               'Ajouts data lake', 'Total'])

    branches = [f for f in wb.sheetnames
                if f not in ('Sommaire', 'Métadonnées', 'Audit', 'Ajouts',
                             'Ajouts non classés')]
    tot_o = tot_v = tot_a = 0
    for b in branches:
        lignes_a = [r for r in wa.iter_rows(min_row=2, values_only=True)
                    if r and r[0] and norm(r[0]) == norm(b)]
        n_o = len(lignes_a)
        n_v = sum(1 for r in lignes_a if not str(r[4]).startswith('NON'))
        n_aj_b = sum(1 for r in waj.iter_rows(min_row=5, values_only=True)
                     if r and r[0] and norm(r[0]) == norm(b))
        ws.append([b, n_o, n_v, n_aj_b, n_o + n_aj_b])
        rr = ws.max_row
        ws.cell(rr, 1).hyperlink = "#'%s'!A1" % b
        ws.cell(rr, 1).font = Font(color='0563C1', underline='single')
        for c in (2, 3, 4, 5):
            ws.cell(rr, c).alignment = Alignment(horizontal='center')
        tot_o += n_o
        tot_v += n_v
        tot_a += n_aj_b
    ws.append(['TOTAL', tot_o, tot_v, tot_a, tot_o + tot_a])
    for c in range(1, 6):
        ws.cell(ws.max_row, c).font = Font(bold=True)
        ws.cell(5, c).font = F_EN
        ws.cell(5, c).fill = EN_TETE
        ws.cell(5, c).alignment = CENTRE
    ws.row_dimensions[5].height = 30
    for c, w in zip('ABCDE', (32, 20, 20, 20, 14)):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = 'A6'
    ws.sheet_view.showGridLines = False

    wb.save(CLASSEUR)
    print('\n%s mis a jour' % CLASSEUR.name)
    print('  colonnes d origine %d | ajouts %d | total %d'
          % (tot_o, tot_a, tot_o + tot_a))


if __name__ == '__main__':
    main()
