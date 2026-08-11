# -*- coding: utf-8 -*-
"""
Met a jour les feuilles de service apres restaurer_confirmees.py :
- Metadonnees : ajoute une ligne par colonne restauree
  (role=Indicateur, origine="Restaurée — source confirmée par l'auteur",
  alignement="confirmé par l'auteur — date non vérifiée")
- Colonnes retirées : ne garde que les 21 colonnes encore effectivement
  retirees (retire les 10 qui ont ete restaurees)
- Sommaire : recalcule les compteurs
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'

from restaurer_confirmees import A_RESTAURER, STATUT   # noqa: E402

F_EN = Font(bold=True, color='FFFFFF', size=10)
EN = PatternFill('solid', fgColor='2F5597')
ORANGE = PatternFill('solid', fgColor='FFE699')


def cle_periode_annee(k):
    import re
    s = str(k)
    m = re.search(r'(19|20)\d{2}', s)
    return int(m.group(0)) if m else None


def main():
    wb = openpyxl.load_workbook(CLASSEUR)

    # ---- Metadonnees : ajouter une ligne par colonne restauree -------------
    wm = wb['Métadonnées']
    for (feuille, titre), inst in A_RESTAURER.items():
        ws = wb[feuille]
        col, axe = None, None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(3, c).value or '').strip() == titre.strip():
                col = c
                break
        if col is None:
            continue
        for cc in range(col, 0, -1):
            v3, v4 = ws.cell(3, cc).value, ws.cell(4, cc).value
            if v3 in ('Trimestre', 'Mois') or v4 in ('Trimestre', 'Mois'):
                axe, type_axe = cc, (v3 if v3 in ('Trimestre', 'Mois') else v4)
                break
        vals = []
        for r in range(5, ws.max_row + 1):
            k, v = ws.cell(r, axe).value, ws.cell(r, col).value
            if k is not None and isinstance(v, (int, float)):
                vals.append((k, v))
        if not vals:
            continue
        debut, fin = vals[0][0], vals[-1][0]
        wm.append([feuille, titre.strip(), 'Indicateur', type_axe,
                  debut, fin, cle_periode_annee(debut), cle_periode_annee(fin),
                  len(vals), inst, 'communiqué par l’auteur de la base',
                  'Bases consolidees (multi-sources)\\BDD SECTORIEL_MENSUEL.csv',
                  titre.strip(), 'Restaurée — source confirmée par l’auteur',
                  STATUT])
        r = wm.max_row
        wm.cell(r, 15).fill = ORANGE
    print('Metadonnees : %d colonnes restaurees ajoutees' % len(A_RESTAURER))

    # ---- Colonnes retirées : ne garder que les 21 encore retirees ----------
    if 'Colonnes retirées' in wb.sheetnames:
        wr = wb['Colonnes retirées']
        a_effacer = []
        for r in range(5, wr.max_row + 1):
            feuille, titre = wr.cell(r, 1).value, str(wr.cell(r, 2).value or '').strip()
            if (feuille, titre) in A_RESTAURER:
                a_effacer.append(r)
        for r in sorted(a_effacer, reverse=True):
            wr.delete_rows(r, 1)
        wr.cell(2, 1).value = (
            'Recherche exhaustive dans les 481 fichiers du data lake (valeur exacte '
            '+ date, tolérance quasi nulle) et tests de transformation (delta, cumul, '
            'arrondi). Aucune correspondance trouvée pour ces colonnes. 10 colonnes '
            'supplémentaires, initialement retirées pour la même raison, ont été '
            'restaurées après confirmation de leur institution productrice par '
            'l’auteur de la base — voir Métadonnées, statut "confirmé par l’auteur — '
            'date non vérifiée".')
        print('Colonnes retirées : %d lignes restantes (10 retirees de cette liste)'
              % (wr.max_row - 4))

    # ---- Sommaire : recalcul -------------------------------------------------
    if 'Sommaire' in wb.sheetnames:
        del wb['Sommaire']
    ws_new = wb.create_sheet('Sommaire', 0)
    ws_new.append(['Branche', 'Colonnes', 'dont cibles', 'dont indicateurs',
                   'infra-annuels', 'alignement non vérifié'])
    from collections import defaultdict
    stats = defaultdict(lambda: [0, 0, 0, 0, 0])
    for r in wm.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        s = stats[r[0]]
        s[0] += 1
        if r[2] in ('Variable cible', 'Cible alternative'):
            s[1] += 1
        else:
            s[2] += 1
        if r[3] in ('Trimestrielle', 'Mensuelle'):
            s[3] += 1
        if r[14] in ('non vérifié', STATUT):
            s[4] += 1

    branches = [f for f in wb.sheetnames
                if f not in ('Sommaire', 'Métadonnées', 'Audit', 'Ajouts',
                             'Ajouts non classés', 'Lisez-moi', 'Colonnes retirées')]
    tot = [0, 0, 0, 0, 0]
    for b in branches:
        s = stats.get(b, [0, 0, 0, 0, 0])
        ws_new.append([b] + s)
        rr = ws_new.max_row
        ws_new.cell(rr, 1).hyperlink = "#'%s'!A1" % b
        ws_new.cell(rr, 1).font = Font(color='0563C1', underline='single')
        for c in range(2, 7):
            ws_new.cell(rr, c).alignment = Alignment(horizontal='center')
        if s[4]:
            ws_new.cell(rr, 6).fill = PatternFill('solid', fgColor='FFEB9C')
        tot = [a + bb for a, bb in zip(tot, s)]
    ws_new.append(['TOTAL'] + tot)
    for c in range(1, 7):
        ws_new.cell(ws_new.max_row, c).font = Font(bold=True)
    for c, w in enumerate((32, 12, 13, 16, 15, 22), start=1):
        ws_new.column_dimensions[get_column_letter(c)].width = w
        cel = ws_new.cell(1, c)
        cel.font = F_EN
        cel.fill = EN
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_new.row_dimensions[1].height = 34
    ws_new.freeze_panes = 'A2'
    ws_new.auto_filter.ref = 'A1:F%d' % ws_new.max_row
    ws_new.sheet_view.showGridLines = False
    print('Sommaire recalcule : %d colonnes au total' % tot[0])

    wb.save(CLASSEUR)
    print('\n%s enregistre.' % CLASSEUR.name)


if __name__ == '__main__':
    main()
