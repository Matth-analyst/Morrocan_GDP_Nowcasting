# -*- coding: utf-8 -*-
"""
Uniformise la signalisation visuelle de la fiabilite des donnees :
- 4 statuts possibles (colonne 15 de Metadonnees), 4 couleurs fixes, une seule fois.
- Retire les incoherences accumulees (orange pose au fil des scripts successifs,
  jaune de "Colonnes retirees" reintroduit par erreur, etc.)
- Ajoute une legende explicite dans Sommaire.

Statuts -> couleur :
  'oui'                                          -> aucune couleur (donnee normale)
  'oui — réaligné'                                -> bleu (DDEBF7) : date corrigee pendant l'audit
  'confirmé par l’auteur — date non vérifiée'     -> orange (FFE699) : source confirmee par
                                                      l'auteur de BDD SECTORIEL, alignement
                                                      des dates non verifiable faute de fichier
                                                      primaire retrouve
  IPAI reconstruit (note specifique)              -> gris (D9D9D9) : indice recalcule par
                                                      chainage de variations trimestrielles,
                                                      pas une serie brute d'origine
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
CLASSEUR = ROOT / 'Etude_sectorielle_Maroc_2_complete.xlsx'

BLANC = PatternFill(fill_type=None)
BLEU = PatternFill('solid', fgColor='DDEBF7')
ORANGE = PatternFill('solid', fgColor='FFE699')
GRIS = PatternFill('solid', fgColor='D9D9D9')

STATUT_CONFIRME = 'confirmé par l’auteur — date non vérifiée'


def couleur_pour(statut):
    if statut is None:
        return BLANC
    s = str(statut)
    if s.startswith(STATUT_CONFIRME):
        return ORANGE
    if 'réaligné' in s:
        return BLEU
    if 'reconstruit' in s.lower() or s.startswith('base interne'):
        return GRIS
    return BLANC


def trouver_titre_col(ws, titre):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(3, c).value or '').strip() == str(titre).strip():
            return c
    return None


def main():
    wb = openpyxl.load_workbook(CLASSEUR)
    wm = wb['Métadonnées']

    # 1) Nettoyer toute couleur existante sur la colonne "Alignement vérifié"
    #    et la recolorer selon le statut réel, un seul passage cohérent.
    n_par_statut = {}
    for r in range(2, wm.max_row + 1):
        feuille = wm.cell(r, 1).value
        titre = wm.cell(r, 2).value
        statut = wm.cell(r, 15).value
        if not feuille or not titre:
            continue
        fill = couleur_pour(statut)
        wm.cell(r, 15).fill = fill
        n_par_statut[statut] = n_par_statut.get(statut, 0) + 1

        # 2) Reporter la meme couleur sur la cellule-titre (ligne 3) de la feuille
        #    de donnees correspondante, en ecrasant toute couleur precedente
        #    posee par un script anterieur (orange systematique, etc.)
        if feuille in wb.sheetnames:
            ws = wb[feuille]
            c = trouver_titre_col(ws, titre)
            if c:
                ws.cell(3, c).fill = fill

    print('Repartition des statuts :')
    for k, v in sorted(n_par_statut.items(), key=lambda x: -x[1]):
        print('  %-70s %d' % (str(k)[:70], v))

    # 3) Colonnes retirées : retirer toute couleur jaune/orange residuelle,
    #    ne garder qu'un simple texte explicatif (pas de "code couleur" a une
    #    seule ligne qui n'a plus de sens depuis les restaurations).
    if 'Colonnes retirées' in wb.sheetnames:
        wr = wb['Colonnes retirées']
        for row in wr.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb not in (None, '00000000'):
                    if cell.row > 1:
                        cell.fill = BLANC

    # 4) Sommaire : renommer la colonne ambigue, retirer le jaune "alerte" par
    #    ligne (qui finissait par colorer presque toutes les branches, donc
    #    n'alertait plus rien), et ajouter une legende claire.
    ws_s = wb['Sommaire']
    ws_s.cell(1, 6).value = 'confirmé auteur (non vérifiable)'
    for r in range(2, ws_s.max_row + 1):
        ws_s.cell(r, 6).fill = BLANC

    # Légende, a partir de la colonne H (8), a cote du tableau
    leg_col = 8
    entetes = [
        ('Légende des couleurs', None, Font(bold=True, size=11)),
        ('Vérifiée (valeur + date confirmées dans la data lake)', BLANC, None),
        ('Vérifiée — réalignée (date corrigée pendant l’audit)', BLEU, None),
        ('Confirmée par l’auteur — alignement non vérifiable', ORANGE, None),
        ('Reconstruite (chaînage / recalcul, pas une série brute)', GRIS, None),
    ]
    for i, (texte, fill, font) in enumerate(entetes):
        cell = ws_s.cell(2 + i, leg_col)
        cell.value = texte
        if fill:
            ws_s.cell(2 + i, leg_col - 1).fill = fill
        if font:
            cell.font = font
        cell.alignment = Alignment(vertical='center')
    ws_s.column_dimensions[get_column_letter(leg_col - 1)].width = 4
    ws_s.column_dimensions[get_column_letter(leg_col)].width = 58

    wb.save(CLASSEUR)
    print('\nNettoyage termine, %s enregistre.' % CLASSEUR.name)


if __name__ == '__main__':
    main()
